"""
CPMS Evaluation Platform – main application
Run locally: python app.py
"""
import os
import json
from datetime import datetime, timezone
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for, flash,
    jsonify, send_from_directory, abort,
)
from werkzeug.utils import secure_filename
from flask_login import (
    LoginManager, login_user, logout_user, login_required, current_user,
)

from config import Config
from models import (
    db, ScoringLevel, Area, Category, TestCase, TestSuite, Project, Vendor,
    TestResult, Evidence, AuditLog, VendorQuestion,
    VendorComment, VendorDocument,
    User, AllowedEmail,
)
from import_excel import import_excel, seed_scoring_levels, generate_test_id
from sqlalchemy import text


def _migrate_schema():
    """Safely add/drop columns introduced after initial deployment (idempotent)."""
    add_stmts = [
        'ALTER TABLE project ADD COLUMN test_suite_id INTEGER REFERENCES test_suite(id)',
        'ALTER TABLE area ADD COLUMN test_suite_id INTEGER REFERENCES test_suite(id)',
    ]
    drop_stmts = [
        'ALTER TABLE evidence DROP COLUMN status',
    ]
    create_stmts = [
        '''CREATE TABLE IF NOT EXISTS vendor_question (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL REFERENCES project(id),
            vendor_id INTEGER REFERENCES vendor(id),
            test_result_id INTEGER REFERENCES test_result(id),
            area_id INTEGER REFERENCES area(id),
            category_id INTEGER REFERENCES category(id),
            question_text TEXT NOT NULL,
            vendor_response TEXT,
            created_at DATETIME,
            responded_at DATETIME,
            created_by VARCHAR(100) DEFAULT 'evaluator',
            status VARCHAR(30) DEFAULT 'Open'
        )''',
        '''CREATE TABLE IF NOT EXISTS vendor_comment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor_id INTEGER NOT NULL REFERENCES vendor(id) ON DELETE CASCADE,
            title VARCHAR(300),
            body TEXT NOT NULL,
            created_at DATETIME,
            updated_at DATETIME,
            created_by VARCHAR(100) DEFAULT 'local_user'
        )''',
        '''CREATE TABLE IF NOT EXISTS vendor_document (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor_id INTEGER NOT NULL REFERENCES vendor(id) ON DELETE CASCADE,
            doc_type VARCHAR(20) DEFAULT 'file',
            filename VARCHAR(300),
            filepath VARCHAR(500),
            url VARCHAR(500),
            description VARCHAR(500),
            uploaded_at DATETIME,
            uploaded_by VARCHAR(100) DEFAULT 'local_user'
        )''',
    ]
    with db.engine.connect() as conn:
        for stmt in create_stmts:
            try:
                conn.execute(text(stmt))
                conn.commit()
            except Exception:
                pass
        for stmt in add_stmts + drop_stmts:
            try:
                conn.execute(text(stmt))
                conn.commit()
            except Exception:
                pass  # column already exists or already dropped – ignore


def create_app():
    app = Flask(__name__)
    app.config.from_object(Config)

    # ensure folders exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(os.path.join(app.instance_path), exist_ok=True)

    db.init_app(app)

    # Flask-Login setup
    login_manager = LoginManager()
    login_manager.login_view = 'login'
    login_manager.login_message = 'Please log in to access the platform.'
    login_manager.login_message_category = 'warning'
    login_manager.init_app(app)

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    with app.app_context():
        db.create_all()
        _migrate_schema()
        seed_scoring_levels()
        _seed_admin()

    return app


def _seed_admin():
    """Create the default admin account if no admin exists yet."""
    if not User.query.filter_by(is_admin=True).first():
        admin_email = os.environ.get('CPMS_ADMIN_EMAIL', 'admin@cpms.local')
        admin_pass  = os.environ.get('CPMS_ADMIN_PASSWORD', 'admin')
        admin = User(
            email=admin_email,
            name='Administrator',
            is_admin=True,
        )
        admin.set_password(admin_pass)
        db.session.add(admin)
        # Also whitelist the admin email
        if not AllowedEmail.query.filter_by(email=admin_email).first():
            db.session.add(AllowedEmail(email=admin_email, added_by='system'))
        db.session.commit()
        print(f'  ✔ Default admin created: {admin_email} / {admin_pass}')
        print(f'    ⚠ Change this password immediately!')


app = create_app()

import re
_NUM_PREFIX_RE = re.compile(r'^\d+(\.\d+)*\.?\s+')

def _strip_number_prefixes():
    """One-time migration: remove leading number prefixes from Area / Category / TestCase.subcategory names."""
    changed = False
    for area in Area.query.all():
        cleaned = _NUM_PREFIX_RE.sub('', area.name)
        if cleaned != area.name:
            area.name = cleaned
            changed = True
    for cat in Category.query.all():
        cleaned = _NUM_PREFIX_RE.sub('', cat.name)
        if cleaned != cat.name:
            cat.name = cleaned
            changed = True
    for tc in TestCase.query.filter(TestCase.subcategory.isnot(None)).all():
        cleaned = _NUM_PREFIX_RE.sub('', tc.subcategory)
        if cleaned != tc.subcategory:
            tc.subcategory = cleaned
            changed = True
    if changed:
        db.session.commit()


def _assign_areas_to_suite():
    """One-time migration: assign orphan Areas to the matching suite template
    based on the test cases they contain.  Falls back to first suite."""
    orphan_areas = Area.query.filter(Area.test_suite_id.is_(None)).all()
    if not orphan_areas:
        return
    first_suite = TestSuite.query.first()
    if not first_suite:
        return
    for area in orphan_areas:
        # Try to find the suite this area's tests actually belong to
        best_suite = None
        for cat in area.categories.all():
            for tc in cat.test_cases.all():
                s = tc.suites.first()
                if s:
                    best_suite = s
                    break
            if best_suite:
                break
        area.test_suite_id = best_suite.id if best_suite else first_suite.id
    db.session.commit()


def _fix_orphan_categories():
    """Idempotent: assign any category with no area_id and no parent_id to an area."""
    orphans = Category.query.filter_by(area_id=None, parent_id=None).all()
    if not orphans:
        return
    for cat in orphans:
        # Find any area matching this suite_type
        area = Area.query.filter_by(suite_type=cat.suite_type).order_by(Area.sort_order).first()
        if not area:
            # Create a fallback area so the category is visible
            max_sort = db.session.query(db.func.max(Area.sort_order)).scalar() or 0
            area = Area(name='Uncategorized', suite_type=cat.suite_type, sort_order=max_sort + 1)
            db.session.add(area)
            db.session.flush()
        cat.area_id = area.id
    db.session.commit()


def _migrate_notes_to_questions():
    """One-time migration: convert existing TestResult.notes into VendorQuestion records."""
    # Only run if there are notes but no questions yet
    has_notes = TestResult.query.filter(TestResult.notes.isnot(None), TestResult.notes != '').first()
    if not has_notes:
        return
    has_questions = VendorQuestion.query.first()
    if has_questions:
        return  # already migrated
    count = 0
    for tr in TestResult.query.filter(TestResult.notes.isnot(None), TestResult.notes != '').all():
        vendor = Vendor.query.get(tr.vendor_id)
        if not vendor:
            continue
        tc = TestCase.query.get(tr.test_case_id)
        area_id = None
        category_id = None
        if tc:
            cat = Category.query.get(tc.category_id)
            if cat:
                category_id = cat.id
                c = cat
                while c:
                    if c.area_id:
                        area_id = c.area_id
                        break
                    c = Category.query.get(c.parent_id) if c.parent_id else None
        qn = VendorQuestion(
            project_id=vendor.project_id,
            vendor_id=vendor.id,
            test_result_id=tr.id,
            area_id=area_id,
            category_id=category_id,
            question_text=tr.notes,
            created_at=tr.updated_at or datetime.now(timezone.utc),
        )
        db.session.add(qn)
        count += 1
    if count:
        db.session.commit()


with app.app_context():
    _strip_number_prefixes()
    _assign_areas_to_suite()
    _fix_orphan_categories()
    _migrate_notes_to_questions()


# ── Helpers ──────────────────────────────────────────────────────────────────
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def _audit(action, entity_type, entity_id, field=None, old=None, new=None, result_id=None):
    entry = AuditLog(
        test_result_id=result_id,
        entity_type=entity_type,
        entity_id=entity_id,
        action=action,
        field_changed=field,
        old_value=str(old) if old is not None else None,
        new_value=str(new) if new is not None else None,
        user=current_user.email if current_user.is_authenticated else 'system',
    )
    db.session.add(entry)


def _audit_changes(entity_type, entity_id, old_vals, new_vals, result_id=None):
    """Compare old_vals dict with new_vals dict and log one audit entry per changed field."""
    for field, old_v in old_vals.items():
        new_v = new_vals.get(field)
        if str(old_v) != str(new_v):
            _audit('updated', entity_type, entity_id,
                   field=field, old=old_v, new=new_v, result_id=result_id)


def _scoring_map():
    """Return dict label→value for scoring levels."""
    return {s.label: s.value for s in ScoringLevel.query.all()}


# ══════════════════════════════════════════════════════════════════════════════
#  AUTHENTICATION ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            login_user(user, remember=request.form.get('remember'))
            next_page = request.args.get('next')
            flash(f'Welcome back, {user.name}!', 'success')
            return redirect(next_page or url_for('dashboard'))
        flash('Invalid email or password.', 'danger')
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        name = request.form.get('name', '').strip()
        password = request.form.get('password', '')
        password2 = request.form.get('password2', '')

        if not email or not name or not password:
            flash('All fields are required.', 'danger')
            return render_template('register.html')
        if password != password2:
            flash('Passwords do not match.', 'danger')
            return render_template('register.html')
        if len(password) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return render_template('register.html')

        # Check whitelist
        allowed = AllowedEmail.query.filter_by(email=email).first()
        if not allowed:
            flash('This email is not authorised to register. Contact your administrator.', 'danger')
            return render_template('register.html')

        if User.query.filter_by(email=email).first():
            flash('An account with this email already exists.', 'danger')
            return render_template('register.html')

        user = User(email=email, name=name)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        login_user(user)
        flash(f'Welcome, {name}! Your account has been created.', 'success')
        return redirect(url_for('dashboard'))
    return render_template('register.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN ROUTES
# ══════════════════════════════════════════════════════════════════════════════

def admin_required(f):
    """Decorator: must be logged-in AND is_admin."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated


@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.order_by(User.created_at.desc()).all()
    allowed = AllowedEmail.query.order_by(AllowedEmail.added_at.desc()).all()
    return render_template('admin_users.html', users=users, allowed=allowed)


@app.route('/admin/whitelist/add', methods=['POST'])
@login_required
@admin_required
def admin_whitelist_add():
    email = request.form.get('email', '').strip().lower()
    if not email:
        flash('Email is required.', 'danger')
        return redirect(url_for('admin_users'))
    if AllowedEmail.query.filter_by(email=email).first():
        flash(f'{email} is already whitelisted.', 'warning')
        return redirect(url_for('admin_users'))
    db.session.add(AllowedEmail(email=email, added_by=current_user.email))
    db.session.commit()
    flash(f'{email} added to whitelist.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/whitelist/<int:entry_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_whitelist_delete(entry_id):
    entry = AllowedEmail.query.get_or_404(entry_id)
    # Don't allow removing an email that has an admin account
    admin_user = User.query.filter_by(email=entry.email, is_admin=True).first()
    if admin_user:
        flash('Cannot remove the admin email from the whitelist.', 'danger')
        return redirect(url_for('admin_users'))
    db.session.delete(entry)
    db.session.commit()
    flash(f'{entry.email} removed from whitelist.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_user_delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_admin:
        flash('Cannot delete the admin account.', 'danger')
        return redirect(url_for('admin_users'))
    db.session.delete(user)
    db.session.commit()
    flash(f'User {user.email} deleted.', 'success')
    return redirect(url_for('admin_users'))


# ══════════════════════════════════════════════════════════════════════════════
#  CONTEXT PROCESSORS
# ══════════════════════════════════════════════════════════════════════════════

@app.context_processor
def inject_sidebar_projects():
    """Inject the list of projects and suites into every template for the sidebar."""
    try:
        projects = Project.query.order_by(Project.name).all()
        suites = TestSuite.query.order_by(TestSuite.name).all()
    except Exception:
        projects = []
        suites = []
    return {'sidebar_projects': projects, 'sidebar_suites': suites}


@app.before_request
def require_login():
    """Redirect unauthenticated users to login for all pages except auth routes."""
    allowed_endpoints = ('login', 'register', 'static')
    if request.endpoint and request.endpoint not in allowed_endpoints and not current_user.is_authenticated:
        return redirect(url_for('login', next=request.url))


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════════════════════════

# ── Dashboard ────────────────────────────────────────────────────────────────
@app.route('/')
def dashboard():
    suites = TestSuite.query.order_by(TestSuite.name).all()
    projects = Project.query.order_by(Project.created_at.desc()).all()
    total_templates = len(suites)
    total_projects = len(projects)
    total_vendors = Vendor.query.count()

    # Active evaluations summary
    active_vendors = db.session.query(Vendor).join(Project).filter(
        Project.status == 'Active'
    ).order_by(Project.name, Vendor.name).all()
    eval_rows = []
    for v in active_vendors:
        total = v.results.count()
        scored = v.results.filter(TestResult.score.isnot(None)).count() if total else 0
        progress = round((scored / total * 100) if total else 0, 1)
        total_ws = db.session.query(db.func.sum(TestResult.weighted_score)).filter(
            TestResult.vendor_id == v.id, TestResult.score.isnot(None)
        ).scalar() or 0
        max_possible = db.session.query(db.func.sum(TestCase.weight)).join(TestResult).filter(
            TestResult.vendor_id == v.id, TestResult.score.isnot(None)
        ).scalar() or 0
        avg_pct = round((total_ws / (max_possible * 5) * 100) if max_possible else 0, 1)
        eval_rows.append({'vendor': v, 'project': v.project, 'total': total,
                          'scored': scored, 'progress': progress, 'avg_pct': avg_pct})

    return render_template('dashboard.html',
                           suites=suites,
                           projects=projects,
                           total_templates=total_templates,
                           total_projects=total_projects,
                           total_vendors=total_vendors,
                           eval_rows=eval_rows)


# ── Import Excel ─────────────────────────────────────────────────────────────
@app.route('/import', methods=['GET', 'POST'])
def import_suite():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file (.xlsx).', 'danger')
            return redirect(url_for('import_suite'))

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        stats = import_excel(filepath)
        if 'error' in stats:
            flash(stats['error'], 'danger')
        else:
            # Auto-create a suite template from the imported tests
            tpl_name = request.form.get('template_name', '').strip()
            if not tpl_name:
                tpl_name = os.path.splitext(filename)[0].replace('_', ' ')
            ts = TestSuite(name=tpl_name, description='Auto-created from Excel import')
            db.session.add(ts)
            db.session.flush()
            # Assign imported areas to this suite
            for area in Area.query.filter(Area.test_suite_id.is_(None)).all():
                area.test_suite_id = ts.id
            for tc in TestCase.query.filter(TestCase.id.in_(stats.get('test_ids', []))).all():
                ts.test_cases.append(tc)
            db.session.commit()
            flash(f"Imported {stats['functional']} functional + {stats['non_functional']} non-functional tests. Suite template \"{tpl_name}\" created with {ts.test_count} test cases.", 'success')
            if stats['warnings']:
                for w in stats['warnings']:
                    flash(w, 'warning')
        return redirect(url_for('dashboard'))

    return render_template('import.html')


# ── Auto-import (for the bundled Excel) ──────────────────────────────────────
@app.route('/import/auto')
def import_auto():
    """One-click import of the CPMS Excel from the parent folder (or Docker mount)."""
    base = os.path.abspath(os.path.dirname(__file__))
    # Check same directory (Docker volume mount) first, then parent folder (local dev)
    candidates = [
        os.path.join(base, 'CPMS_Evaluation_Test_Suite.xlsx'),
        os.path.join(os.path.dirname(base), 'CPMS_Evaluation_Test_Suite.xlsx'),
    ]
    xlsx = next((p for p in candidates if os.path.exists(p)), None)
    if not xlsx:
        flash('CPMS_Evaluation_Test_Suite.xlsx not found.', 'danger')
        return redirect(url_for('import_suite'))

    if TestCase.query.count() > 0:
        flash('Test cases already imported. Delete database to re-import.', 'warning')
        return redirect(url_for('dashboard'))

    stats = import_excel(xlsx)
    if 'error' in stats:
        flash(stats['error'], 'danger')
    else:
        ts = TestSuite(name='CPMS Evaluation Suite', description='Auto-imported from CPMS Excel')
        db.session.add(ts)
        db.session.flush()
        # Assign orphan areas to this new suite
        for area in Area.query.filter(Area.test_suite_id.is_(None)).all():
            area.test_suite_id = ts.id
        for tc in TestCase.query.filter(TestCase.id.in_(stats.get('test_ids', []))).all():
            ts.test_cases.append(tc)
        db.session.commit()
        flash(f"Auto-imported {stats['functional']} functional + {stats['non_functional']} non-functional tests into template \"{ts.name}\".", 'success')
    return redirect(url_for('dashboard'))


# ══════════════════════════════════════════════════════════════════════════════
#  TEST CASE CRUD  (always in the context of a suite template)
# ══════════════════════════════════════════════════════════════════════════════


@app.route('/tests/<int:test_id>')
def test_detail(test_id):
    tc = TestCase.query.get_or_404(test_id)
    suite_id = request.args.get('suite_id', type=int)
    ts = TestSuite.query.get(suite_id) if suite_id else tc.suites.first()
    return render_template('test_detail.html', test=tc, ts=ts)


# ── JSON API for modal editing ──────────────────────────────────────────────
@app.route('/api/tests/<int:test_id>', methods=['GET'])
def api_test_get(test_id):
    """Return test case data as JSON for the edit modal."""
    tc = TestCase.query.get_or_404(test_id)
    cat = tc.category
    if cat.parent_id:
        current_sub_id = cat.id
        parent_cat = Category.query.get(cat.parent_id)
        current_cat_id = parent_cat.id
        current_area_id = parent_cat.area_id
    else:
        current_sub_id = None
        current_cat_id = cat.id
        current_area_id = cat.area_id

    # Resolve suite for area options
    suite_id = request.args.get('suite_id', type=int)
    ts = TestSuite.query.get(suite_id) if suite_id else tc.suites.first()
    if ts:
        areas = Area.query.filter_by(test_suite_id=ts.id, suite_type=tc.suite_type).order_by(Area.sort_order).all()
    else:
        areas = Area.query.filter_by(suite_type=tc.suite_type).order_by(Area.sort_order).all()

    area_list = []
    for a in areas:
        cats = []
        for c in a.categories:
            if not c.parent_id:
                subs = [{'id': s.id, 'name': s.name} for s in c.children]
                cats.append({'id': c.id, 'name': c.name, 'subcategories': subs})
        area_list.append({'id': a.id, 'name': a.name, 'categories': cats})

    return jsonify({
        'id': tc.id,
        'test_id_code': tc.test_id_code,
        'capability': tc.capability or '',
        'test_scenario': tc.test_scenario or '',
        'pass_criteria': tc.pass_criteria or '',
        'evidence_required': tc.evidence_required or '',
        'test_method': tc.test_method or '',
        'priority': tc.priority or 'Should',
        'tier': tc.tier or 'Core',
        'weight': tc.weight,
        'is_mandatory': tc.is_mandatory,
        'suite_type': tc.suite_type or '',
        'current_area_id': current_area_id,
        'current_cat_id': current_cat_id,
        'current_sub_id': current_sub_id,
        'areas': area_list,
    })


@app.route('/api/tests/<int:test_id>', methods=['PUT'])
def api_test_update(test_id):
    """Save test case data from modal form (AJAX)."""
    tc = TestCase.query.get_or_404(test_id)
    data = request.get_json()

    # Snapshot old values for audit
    old_vals = {
        'capability': tc.capability, 'test_scenario': tc.test_scenario,
        'pass_criteria': tc.pass_criteria, 'evidence_required': tc.evidence_required,
        'test_method': tc.test_method, 'priority': tc.priority, 'tier': tc.tier,
        'weight': tc.weight, 'is_mandatory': tc.is_mandatory,
        'category_id': tc.category_id, 'subcategory': tc.subcategory,
    }

    tc.capability = data.get('capability', tc.capability)
    tc.test_scenario = data.get('test_scenario', tc.test_scenario)
    tc.pass_criteria = data.get('pass_criteria', tc.pass_criteria)
    tc.evidence_required = data.get('evidence_required', tc.evidence_required)
    tc.test_method = data.get('test_method', tc.test_method)
    tc.priority = data.get('priority', tc.priority)
    tc.tier = data.get('tier', tc.tier)
    if data.get('weight') is not None:
        tc.weight = float(data['weight'])
    tc.is_mandatory = bool(data.get('is_mandatory'))

    sub_cat_id = data.get('subcategory_id')
    cat_id = data.get('category_id')
    if sub_cat_id:
        tc.category_id = int(sub_cat_id)
        sub = Category.query.get(int(sub_cat_id))
        tc.subcategory = sub.name if sub else None
    elif cat_id:
        tc.category_id = int(cat_id)
        tc.subcategory = None

    new_vals = {
        'capability': tc.capability, 'test_scenario': tc.test_scenario,
        'pass_criteria': tc.pass_criteria, 'evidence_required': tc.evidence_required,
        'test_method': tc.test_method, 'priority': tc.priority, 'tier': tc.tier,
        'weight': tc.weight, 'is_mandatory': tc.is_mandatory,
        'category_id': tc.category_id, 'subcategory': tc.subcategory,
    }
    _audit_changes('TestCase', tc.id, old_vals, new_vals)
    db.session.commit()
    return jsonify({'ok': True, 'test_id_code': tc.test_id_code})


@app.route('/api/tests', methods=['POST'])
def api_test_create():
    """Create a new test case via JSON (AJAX modal)."""
    data = request.get_json()
    suite_type = data.get('suite_type', 'Functional')
    capability = (data.get('capability') or '').strip()
    if not capability:
        return jsonify({'ok': False, 'error': 'Capability / feature description is required.'}), 400

    test_id_code = generate_test_id(suite_type)

    sub_cat_id = data.get('subcategory_id')
    cat_id = data.get('category_id')
    if sub_cat_id:
        category_id = int(sub_cat_id)
        sub = Category.query.get(int(sub_cat_id))
        subcategory_name = sub.name if sub else None
    elif cat_id:
        category_id = int(cat_id)
        subcategory_name = None
    else:
        return jsonify({'ok': False, 'error': 'Please select a category.'}), 400

    weight = 1.0
    try:
        weight = float(data.get('weight', 1.0))
    except (ValueError, TypeError):
        pass

    priority = data.get('priority', 'Should')
    max_sort = db.session.query(db.func.max(TestCase.sort_order)).filter_by(suite_type=suite_type).scalar() or 0

    tc = TestCase(
        test_id_code=test_id_code,
        tier=data.get('tier', 'Core'),
        category_id=category_id,
        subcategory=subcategory_name,
        capability=capability,
        test_scenario=(data.get('test_scenario') or '').strip() or None,
        pass_criteria=(data.get('pass_criteria') or '').strip() or None,
        evidence_required=(data.get('evidence_required') or '').strip() or None,
        test_method=(data.get('test_method') or '').strip() or None,
        priority=priority,
        weight=weight,
        is_mandatory=(priority.lower() == 'must') or bool(data.get('is_mandatory')),
        suite_type=suite_type,
        sort_order=max_sort + 1,
    )
    db.session.add(tc)
    db.session.flush()

    suite_id = data.get('suite_id')
    target_suite = TestSuite.query.get(int(suite_id)) if suite_id else None
    if target_suite:
        target_suite.test_cases.append(tc)
        for proj in target_suite.projects.all():
            for v in proj.vendors.all():
                existing = TestResult.query.filter_by(vendor_id=v.id, test_case_id=tc.id).first()
                if not existing:
                    db.session.add(TestResult(vendor_id=v.id, test_case_id=tc.id, status='Not Started'))

    _audit('created', 'TestCase', tc.id)
    db.session.commit()
    return jsonify({'ok': True, 'test_id_code': tc.test_id_code, 'id': tc.id})


@app.route('/tests/<int:test_id>/edit', methods=['GET', 'POST'])
def test_edit(test_id):
    tc = TestCase.query.get_or_404(test_id)
    # Resolve suite for scoped areas
    suite_id = request.args.get('suite_id', type=int)
    ts = TestSuite.query.get(suite_id) if suite_id else tc.suites.first()
    if ts:
        areas = Area.query.filter_by(test_suite_id=ts.id, suite_type=tc.suite_type).order_by(Area.sort_order).all()
    else:
        areas = Area.query.filter_by(suite_type=tc.suite_type).order_by(Area.sort_order).all()

    # Determine current area, category, subcategory from the hierarchy
    cat = tc.category
    if cat.parent_id:
        # category_id points to a subcategory
        current_sub_id = cat.id
        parent_cat = Category.query.get(cat.parent_id)
        current_cat_id = parent_cat.id
        current_area_id = parent_cat.area_id
    else:
        current_sub_id = None
        current_cat_id = cat.id
        current_area_id = cat.area_id

    if request.method == 'POST':
        # Snapshot old values for audit
        old_vals = {
            'capability': tc.capability, 'test_scenario': tc.test_scenario,
            'pass_criteria': tc.pass_criteria, 'evidence_required': tc.evidence_required,
            'test_method': tc.test_method, 'priority': tc.priority, 'tier': tc.tier,
            'weight': tc.weight, 'is_mandatory': tc.is_mandatory,
            'category_id': tc.category_id, 'subcategory': tc.subcategory,
        }

        tc.capability = request.form.get('capability', tc.capability)
        tc.test_scenario = request.form.get('test_scenario', tc.test_scenario)
        tc.pass_criteria = request.form.get('pass_criteria', tc.pass_criteria)
        tc.evidence_required = request.form.get('evidence_required', tc.evidence_required)
        tc.test_method = request.form.get('test_method', tc.test_method)
        tc.priority = request.form.get('priority', tc.priority)
        tc.tier = request.form.get('tier', tc.tier)
        new_weight = request.form.get('weight')
        if new_weight:
            tc.weight = float(new_weight)
        tc.is_mandatory = 'is_mandatory' in request.form

        # Handle area / category / subcategory
        sub_cat_id = request.form.get('subcategory_id')
        cat_id = request.form.get('category_id')
        if sub_cat_id:
            tc.category_id = int(sub_cat_id)
            sub = Category.query.get(int(sub_cat_id))
            tc.subcategory = sub.name if sub else None
        elif cat_id:
            tc.category_id = int(cat_id)
            tc.subcategory = None

        new_vals = {
            'capability': tc.capability, 'test_scenario': tc.test_scenario,
            'pass_criteria': tc.pass_criteria, 'evidence_required': tc.evidence_required,
            'test_method': tc.test_method, 'priority': tc.priority, 'tier': tc.tier,
            'weight': tc.weight, 'is_mandatory': tc.is_mandatory,
            'category_id': tc.category_id, 'subcategory': tc.subcategory,
        }
        _audit_changes('TestCase', tc.id, old_vals, new_vals)
        db.session.commit()
        flash('Test case updated.', 'success')
        next_url = request.form.get('next') or request.args.get('next')
        if next_url:
            return redirect(next_url)
        suite_id_redir = request.args.get('suite_id', type=int)
        return redirect(url_for('test_detail', test_id=tc.id, suite_id=suite_id_redir))

    # Determine where to go back to after editing
    back_url = request.args.get('next') or request.referrer or url_for('test_detail', test_id=tc.id, suite_id=suite_id)
    return render_template('test_edit.html', test=tc,
                           areas=areas,
                           current_area_id=current_area_id,
                           current_cat_id=current_cat_id,
                           current_sub_id=current_sub_id,
                           ts=ts,
                           back_url=back_url)


@app.route('/tests/new', methods=['GET', 'POST'])
def test_new():
    suite = request.args.get('suite', 'Functional')
    suite_id = request.args.get('suite_id', type=int)
    ts = TestSuite.query.get(suite_id) if suite_id else None
    if ts:
        areas_func = Area.query.filter_by(test_suite_id=ts.id, suite_type='Functional').order_by(Area.sort_order).all()
        areas_nf = Area.query.filter_by(test_suite_id=ts.id, suite_type='Non-Functional').order_by(Area.sort_order).all()
    else:
        areas_func = Area.query.filter_by(suite_type='Functional').order_by(Area.sort_order).all()
        areas_nf = Area.query.filter_by(suite_type='Non-Functional').order_by(Area.sort_order).all()

    # Pre-selection support: ?precat_id=<category_or_subcategory_id>
    pre_area_id = None
    pre_cat_id = None
    pre_sub_id = None
    precat_id = request.args.get('precat_id', type=int)
    if precat_id:
        cat = Category.query.get(precat_id)
        if cat:
            suite = cat.suite_type  # match suite to category
            if cat.parent_id:
                # It's a subcategory
                pre_sub_id = cat.id
                parent = Category.query.get(cat.parent_id)
                if parent:
                    pre_cat_id = parent.id
                    pre_area_id = parent.area_id
            else:
                pre_cat_id = cat.id
                pre_area_id = cat.area_id

    if request.method == 'POST':
        suite_type = request.form.get('suite_type', 'Functional')
        capability = request.form.get('capability', '').strip()
        if not capability:
            flash('Capability / feature description is required.', 'danger')
            return redirect(url_for('test_new', suite=suite_type))

        # Auto-generate test ID
        test_id_code = generate_test_id(suite_type)

        # Handle area / category / subcategory
        sub_cat_id = request.form.get('subcategory_id')
        cat_id = request.form.get('category_id')
        if sub_cat_id:
            category_id = int(sub_cat_id)
            sub = Category.query.get(int(sub_cat_id))
            subcategory_name = sub.name if sub else None
        elif cat_id:
            category_id = int(cat_id)
            subcategory_name = None
        else:
            flash('Please select a category.', 'danger')
            return redirect(url_for('test_new', suite=suite_type))

        weight = 1.0
        try:
            weight = float(request.form.get('weight', 1.0))
        except (ValueError, TypeError):
            pass

        priority = request.form.get('priority', 'Should')
        max_sort = db.session.query(db.func.max(TestCase.sort_order)).filter_by(suite_type=suite_type).scalar() or 0

        tc = TestCase(
            test_id_code=test_id_code,
            tier=request.form.get('tier', 'Core'),
            category_id=category_id,
            subcategory=subcategory_name,
            capability=capability,
            test_scenario=request.form.get('test_scenario', '').strip() or None,
            pass_criteria=request.form.get('pass_criteria', '').strip() or None,
            evidence_required=request.form.get('evidence_required', '').strip() or None,
            test_method=request.form.get('test_method', '').strip() or None,
            priority=priority,
            weight=weight,
            is_mandatory=(priority.lower() == 'must') or ('is_mandatory' in request.form),
            suite_type=suite_type,
            sort_order=max_sort + 1,
        )
        db.session.add(tc)
        db.session.flush()  # get tc.id

        # Auto-add to suite template if one was specified
        target_suite = TestSuite.query.get(suite_id) if suite_id else None
        if target_suite:
            target_suite.test_cases.append(tc)

        # Auto-create TestResult entries for vendors whose project uses this suite
        if target_suite:
            for proj in target_suite.projects.all():
                for v in proj.vendors.all():
                    existing = TestResult.query.filter_by(vendor_id=v.id, test_case_id=tc.id).first()
                    if not existing:
                        tr = TestResult(vendor_id=v.id, test_case_id=tc.id, status='Not Started')
                        db.session.add(tr)

        _audit('created', 'TestCase', tc.id)
        db.session.commit()
        flash(f'Test case {test_id_code} created.', 'success')
        if target_suite:
            return redirect(url_for('suite_detail', suite_id=target_suite.id, suite=suite_type))
        return redirect(url_for('test_detail', test_id=tc.id))

    ts = TestSuite.query.get(suite_id) if suite_id else None
    return render_template('test_new.html', suite=suite,
                           areas_func=areas_func,
                           areas_nf=areas_nf,
                           pre_area_id=pre_area_id,
                           pre_cat_id=pre_cat_id,
                           pre_sub_id=pre_sub_id,
                           ts=ts,
                           suite_id=suite_id)


@app.route('/tests/<int:test_id>/delete', methods=['POST'])
def test_delete(test_id):
    tc = TestCase.query.get_or_404(test_id)
    suite = tc.suite_type
    tid = tc.test_id_code
    suite_id = request.form.get('suite_id', type=int)

    # Delete associated results, evidence, and audit entries
    for tr in tc.results.all():
        Evidence.query.filter_by(test_result_id=tr.id).delete()
        AuditLog.query.filter_by(test_result_id=tr.id).delete()
        db.session.delete(tr)

    _audit('deleted', 'TestCase', tc.id)
    db.session.delete(tc)
    db.session.commit()
    flash(f'Test case {tid} deleted.', 'success')
    if suite_id:
        return redirect(url_for('suite_detail', suite_id=suite_id, suite=suite))
    return redirect(url_for('suite_list'))


# ══════════════════════════════════════════════════════════════════════════════
#  AREAS  (scoped to a suite template)
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/areas')
def area_list():
    """Backward-compat redirect – sends to first suite's categories."""
    first_suite = TestSuite.query.first()
    if first_suite:
        return redirect(url_for('category_list', suite_id=first_suite.id))
    return redirect(url_for('suite_list'))


@app.route('/suites/<int:suite_id>/areas/new', methods=['GET', 'POST'])
def area_new(suite_id):
    ts = TestSuite.query.get_or_404(suite_id)
    suite = request.args.get('suite', 'Functional')
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        suite_type = request.form.get('suite_type', 'Functional')
        if not name:
            flash('Area name is required.', 'danger')
            return redirect(url_for('area_new', suite_id=suite_id, suite=suite_type))
        max_sort = db.session.query(db.func.max(Area.sort_order)).filter_by(test_suite_id=suite_id, suite_type=suite_type).scalar() or 0
        area = Area(name=name, suite_type=suite_type, test_suite_id=suite_id, sort_order=max_sort + 1)
        db.session.add(area)
        _audit('created', 'Area', 0)
        db.session.commit()
        flash(f'Area "{name}" created.', 'success')
        return redirect(url_for('category_list', suite_id=suite_id, suite=suite_type))
    return render_template('area_form.html', area=None, suite=suite, ts=ts)


@app.route('/areas/<int:area_id>/edit', methods=['GET', 'POST'])
def area_edit(area_id):
    area = Area.query.get_or_404(area_id)
    ts = TestSuite.query.get(area.test_suite_id) if area.test_suite_id else None
    if request.method == 'POST':
        old_name = area.name
        area.name = request.form.get('name', area.name).strip()
        _audit('updated', 'Area', area.id, 'name', old_name, area.name)
        db.session.commit()
        flash('Area updated.', 'success')
        if ts:
            return redirect(url_for('category_list', suite_id=ts.id, suite=area.suite_type))
        return redirect(url_for('suite_list'))
    return render_template('area_form.html', area=area, suite=area.suite_type, ts=ts)


@app.route('/areas/<int:area_id>/delete', methods=['POST'])
def area_delete(area_id):
    area = Area.query.get_or_404(area_id)
    suite_id = area.test_suite_id
    suite = area.suite_type
    # Reassign categories to no area
    for cat in area.categories:
        cat.area_id = None
    db.session.delete(area)
    db.session.commit()
    flash('Area deleted. Categories have been unassigned.', 'success')
    if suite_id:
        return redirect(url_for('category_list', suite_id=suite_id, suite=suite))
    return redirect(url_for('suite_list'))


# ══════════════════════════════════════════════════════════════════════════════
#  CATEGORIES  (scoped to a suite template)
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/categories')
def category_list_legacy():
    """Backward-compat redirect."""
    first_suite = TestSuite.query.first()
    if first_suite:
        return redirect(url_for('category_list', suite_id=first_suite.id,
                                suite=request.args.get('suite', 'Functional')))
    return redirect(url_for('suite_list'))


@app.route('/suites/<int:suite_id>/categories')
def category_list(suite_id):
    ts = TestSuite.query.get_or_404(suite_id)
    suite = request.args.get('suite', 'Functional')
    areas = Area.query.filter_by(test_suite_id=suite_id, suite_type=suite).order_by(Area.sort_order).all()
    return render_template('categories.html', areas=areas, suite=suite, ts=ts)


@app.route('/suites/<int:suite_id>/categories/new', methods=['GET', 'POST'])
def category_new(suite_id):
    ts = TestSuite.query.get_or_404(suite_id)
    suite = request.args.get('suite', 'Functional')
    areas = Area.query.filter_by(test_suite_id=suite_id, suite_type=suite).order_by(Area.sort_order).all()
    parents = Category.query.filter(
        Category.area_id.in_([a.id for a in areas]),
        Category.parent_id.is_(None),
    ).order_by(Category.sort_order).all()

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        suite_type = request.form.get('suite_type', 'Functional')
        area_id = request.form.get('area_id') or None
        parent_id = request.form.get('parent_id') or None
        weight_mult = float(request.form.get('weight_multiplier', 1.0))
        if not name:
            flash('Category name is required.', 'danger')
            return redirect(url_for('category_new', suite_id=suite_id, suite=suite_type))

        max_sort = db.session.query(db.func.max(Category.sort_order)).filter_by(suite_type=suite_type).scalar() or 0
        cat = Category(
            name=name, suite_type=suite_type,
            area_id=int(area_id) if area_id else None,
            parent_id=int(parent_id) if parent_id else None,
            weight_multiplier=weight_mult, sort_order=max_sort + 1,
        )
        db.session.add(cat)
        _audit('created', 'Category', 0)
        db.session.commit()
        flash(f'Category "{name}" created.', 'success')
        return redirect(url_for('category_list', suite_id=suite_id, suite=suite_type))

    return render_template('category_form.html', category=None, areas=areas, parents=parents, suite=suite, ts=ts)


@app.route('/categories/<int:cat_id>/edit', methods=['GET', 'POST'])
def category_edit(cat_id):
    cat = Category.query.get_or_404(cat_id)
    # Find the suite this category belongs to via its area
    area = Area.query.get(cat.area_id) if cat.area_id else None
    ts = TestSuite.query.get(area.test_suite_id) if area and area.test_suite_id else None
    suite_id = ts.id if ts else None
    areas = Area.query.filter_by(test_suite_id=suite_id, suite_type=cat.suite_type).order_by(Area.sort_order).all() if suite_id else []
    parents = Category.query.filter(
        Category.area_id.in_([a.id for a in areas]),
        Category.parent_id.is_(None),
        Category.id != cat.id,
    ).order_by(Category.sort_order).all()

    if request.method == 'POST':
        old_name = cat.name
        cat.name = request.form.get('name', cat.name).strip()
        cat.weight_multiplier = float(request.form.get('weight_multiplier', cat.weight_multiplier))
        area_id = request.form.get('area_id') or None
        cat.area_id = int(area_id) if area_id else None
        parent_id = request.form.get('parent_id') or None
        cat.parent_id = int(parent_id) if parent_id else None
        _audit('updated', 'Category', cat.id, 'name', old_name, cat.name)
        db.session.commit()
        flash('Category updated.', 'success')
        if suite_id:
            return redirect(url_for('category_list', suite_id=suite_id, suite=cat.suite_type))
        return redirect(url_for('suite_list'))

    return render_template('category_form.html', category=cat, areas=areas, parents=parents, suite=cat.suite_type, ts=ts)


@app.route('/categories/<int:cat_id>/delete', methods=['POST'])
def category_delete(cat_id):
    cat = Category.query.get_or_404(cat_id)
    suite_type = cat.suite_type
    # Find suite for redirect — walk up parent chain for subcategories
    area = None
    lookup = cat
    while lookup and not area:
        if lookup.area_id:
            area = Area.query.get(lookup.area_id)
        elif lookup.parent_id:
            lookup = Category.query.get(lookup.parent_id)
        else:
            break
    suite_id = area.test_suite_id if area else None

    def _delete_tests_for_category(c):
        """Delete all test cases (and their results/evidence/audit/suite-links) for a category."""
        for tc in list(c.test_cases):
            # Remove suite links
            db.session.execute(db.text('DELETE FROM suite_tests WHERE test_case_id = :tid'), {'tid': tc.id})
            # Remove results cascade
            for tr in TestResult.query.filter_by(test_case_id=tc.id).all():
                Evidence.query.filter_by(test_result_id=tr.id).delete()
                AuditLog.query.filter_by(test_result_id=tr.id).delete()
                db.session.delete(tr)
            db.session.delete(tc)

    # Recursively delete children first
    def _delete_category_tree(c):
        for child in list(c.children):
            _delete_category_tree(child)
        _delete_tests_for_category(c)
        db.session.delete(c)

    _delete_category_tree(cat)
    db.session.commit()

    # If AJAX request, return JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return {'ok': True, 'message': 'Category and its test cases deleted.'}

    flash('Category and its test cases deleted.', 'success')
    if suite_id:
        return redirect(url_for('category_list', suite_id=suite_id, suite=suite_type))
    return redirect(url_for('suite_list'))


# ══════════════════════════════════════════════════════════════════════════════
#  PROJECTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/projects')
def project_list():
    projects = Project.query.order_by(Project.created_at.desc()).all()
    return render_template('projects.html', projects=projects)


@app.route('/projects/new', methods=['GET', 'POST'])
def project_new():
    suites = TestSuite.query.order_by(TestSuite.name).all()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        desc = request.form.get('description', '').strip()
        if not name:
            flash('Project name is required.', 'danger')
            return redirect(url_for('project_new'))

        suite_mode = request.form.get('suite_mode', 'new')
        if suite_mode == 'existing':
            suite_id = request.form.get('suite_id', type=int)
            ts = TestSuite.query.get(suite_id) if suite_id else None
            if not ts:
                flash('Please select a valid suite template.', 'danger')
                return redirect(url_for('project_new'))
        else:
            suite_name = request.form.get('suite_name', '').strip() or name
            ts = TestSuite(name=suite_name, description=request.form.get('suite_description', '').strip() or None)
            db.session.add(ts)
            db.session.flush()

        p = Project(name=name, description=desc, test_suite_id=ts.id)
        db.session.add(p)
        db.session.commit()
        flash(f'Project "{name}" created using suite template "{ts.name}" ({ts.test_count} test cases).', 'success')
        return redirect(url_for('project_detail', project_id=p.id))
    return render_template('project_form.html', project=None, suites=suites)


@app.route('/projects/<int:project_id>')
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)
    vendors = project.vendors.all()

    # Compute stats for each vendor (progress, avg score, etc.)
    vendor_stats = []
    for v in vendors:
        total = v.results.count()
        scored = v.results.filter(TestResult.score.isnot(None)).count() if total else 0
        total_ws = db.session.query(db.func.sum(TestResult.weighted_score)).filter(
            TestResult.vendor_id == v.id, TestResult.score.isnot(None)
        ).scalar() or 0
        max_possible = db.session.query(db.func.sum(TestCase.weight)).join(TestResult).filter(
            TestResult.vendor_id == v.id, TestResult.score.isnot(None)
        ).scalar() or 0
        max_score = 5  # top of the scoring scale
        avg_pct = round((total_ws / (max_possible * max_score) * 100) if max_possible else 0, 1)
        progress = round((scored / total * 100) if total else 0, 1)

        vendor_stats.append({
            'vendor': v,
            'total': total,
            'scored': scored,
            'avg_pct': avg_pct,
            'progress': progress,
        })

    return render_template('project_detail.html', project=project, vendors=vendors, vendor_stats=vendor_stats)


@app.route('/projects/<int:project_id>/edit', methods=['GET', 'POST'])
def project_edit(project_id):
    project = Project.query.get_or_404(project_id)
    suites = TestSuite.query.order_by(TestSuite.name).all()
    if request.method == 'POST':
        project.name = request.form.get('name', project.name).strip()
        project.description = request.form.get('description', project.description)
        project.status = request.form.get('status', project.status)
        new_suite_id = request.form.get('suite_id', type=int)
        if new_suite_id and new_suite_id != project.test_suite_id:
            project.test_suite_id = new_suite_id
            flash('Suite template updated. New vendors added to this project will use the new template.', 'info')
        db.session.commit()
        flash('Project updated.', 'success')
        return redirect(url_for('project_detail', project_id=project.id))
    return render_template('project_form.html', project=project, suites=suites)


@app.route('/projects/<int:project_id>/delete', methods=['POST'])
def project_delete(project_id):
    project = Project.query.get_or_404(project_id)
    name = project.name

    # Delete all child data: questions → evidence → audit → results → vendors
    VendorQuestion.query.filter_by(project_id=project_id).delete()
    for vendor in project.vendors.all():
        for tr in vendor.results.all():
            Evidence.query.filter_by(test_result_id=tr.id).delete()
            AuditLog.query.filter_by(test_result_id=tr.id).delete()
            db.session.delete(tr)
        db.session.delete(vendor)

    # Detach & delete the project's test suite if it belongs only to this project
    ts_id = project.test_suite_id
    project.test_suite_id = None
    db.session.flush()
    if ts_id:
        ts = TestSuite.query.get(ts_id)
        if ts and ts.projects.count() == 0:
            db.session.delete(ts)

    _audit('deleted', 'Project', project.id)
    db.session.delete(project)
    db.session.commit()
    flash(f'Project "{name}" and all associated data deleted.', 'success')
    return redirect(url_for('project_list'))


# ══════════════════════════════════════════════════════════════════════════════
#  TEST SUITES  (per-project test case collections)
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/suites')
def suite_list():
    suites = TestSuite.query.order_by(TestSuite.name).all()
    return render_template('suites.html', suites=suites)


@app.route('/suites/new', methods=['GET', 'POST'])
def suite_new():
    existing_suites = TestSuite.query.order_by(TestSuite.name).all()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        desc = request.form.get('description', '').strip() or None
        if not name:
            flash('Suite name is required.', 'danger')
            return redirect(url_for('suite_new'))
        ts = TestSuite(name=name, description=desc)
        db.session.add(ts)
        db.session.flush()
        clone_id = request.form.get('clone_from', type=int)
        if clone_id:
            source = TestSuite.query.get(clone_id)
            if source:
                for tc in source.test_cases.order_by(TestCase.sort_order).all():
                    ts.test_cases.append(tc)
        db.session.commit()
        flash(f'Suite template "{name}" created with {ts.test_count} test cases.', 'success')
        return redirect(url_for('suite_detail', suite_id=ts.id))
    return render_template('suite_form.html', ts=None, existing_suites=existing_suites)


@app.route('/suites/<int:suite_id>/edit', methods=['GET', 'POST'])
def suite_edit(suite_id):
    ts = TestSuite.query.get_or_404(suite_id)
    if request.method == 'POST':
        ts.name = request.form.get('name', ts.name).strip()
        ts.description = request.form.get('description', '').strip() or None
        db.session.commit()
        flash('Suite template updated.', 'success')
        return redirect(url_for('suite_detail', suite_id=ts.id))
    return render_template('suite_form.html', ts=ts)


@app.route('/suites/<int:suite_id>/delete', methods=['POST'])
def suite_delete(suite_id):
    ts = TestSuite.query.get_or_404(suite_id)
    if ts.projects.count() > 0:
        flash('Cannot delete a suite template that is still used by projects.', 'danger')
        return redirect(url_for('suite_detail', suite_id=suite_id))
    name = ts.name  # capture before delete so it survives commit expiry
    db.session.delete(ts)
    db.session.commit()
    flash(f'Suite template "{name}" deleted.', 'success')
    return redirect(url_for('suite_list'))


@app.route('/suites/<int:suite_id>')
def suite_detail(suite_id):
    ts = TestSuite.query.get_or_404(suite_id)
    suite_type = request.args.get('suite', 'Functional')
    tier = request.args.get('tier', '')
    priority = request.args.get('priority', '')
    area_id = request.args.get('area', '', type=str)
    search = request.args.get('search', '')

    cat_id = request.args.get('cat', '', type=str)
    subcat_id = request.args.get('subcat', '', type=str)

    q = ts.test_cases.filter(TestCase.suite_type == suite_type)
    if tier:
        q = q.filter(TestCase.tier == tier)
    if priority:
        q = q.filter(TestCase.priority == priority)
    if subcat_id:
        # Filter by specific subcategory only
        q = q.filter(TestCase.category_id == int(subcat_id))
    elif cat_id:
        # Filter by specific category (include its children)
        target_ids = [int(cat_id)]
        child_ids = [c.id for c in Category.query.filter_by(parent_id=int(cat_id)).all()]
        target_ids.extend(child_ids)
        q = q.filter(TestCase.category_id.in_(target_ids))
    elif area_id:
        area_cat_ids = [c.id for c in Category.query.filter_by(area_id=int(area_id)).all()]
        # Include subcategories
        sub_ids = [s.id for s in Category.query.filter(Category.parent_id.in_(area_cat_ids)).all()]
        all_cat_ids = area_cat_ids + sub_ids
        if all_cat_ids:
            q = q.filter(TestCase.category_id.in_(all_cat_ids))
    if search:
        q = q.filter(db.or_(
            TestCase.test_id_code.ilike(f'%{search}%'),
            TestCase.capability.ilike(f'%{search}%'),
            TestCase.test_scenario.ilike(f'%{search}%'),
        ))

    tests = q.order_by(TestCase.sort_order).all()
    areas = Area.query.filter_by(test_suite_id=suite_id, suite_type=suite_type).order_by(Area.sort_order).all()
    # Build category & subcategory lists for filter dropdowns + index map
    categories = []
    subcategories = []
    cat_index = {}   # {category_id: '1.2.3' style index}
    for ai, a in enumerate(areas, 1):
        cat_index[f'area_{a.id}'] = str(ai)
        ci = 0
        for c in Category.query.filter_by(area_id=a.id, parent_id=None).order_by(Category.sort_order).all():
            ci += 1
            categories.append(c)
            cat_index[c.id] = f'{ai}.{ci}'
            for si, sc in enumerate(c.children.order_by(Category.sort_order).all(), 1):
                subcategories.append(sc)
                cat_index[sc.id] = f'{ai}.{ci}.{si}'
    return render_template('suite_detail.html', ts=ts, tests=tests,
                           suite_type=suite_type, areas=areas,
                           categories=categories, subcategories=subcategories,
                           cat_index=cat_index)


@app.route('/suites/<int:suite_id>/add_test', methods=['POST'])
def suite_add_test(suite_id):
    ts = TestSuite.query.get_or_404(suite_id)
    tc_id = request.form.get('test_case_id', type=int)
    if tc_id:
        tc = TestCase.query.get(tc_id)
        if tc and tc not in ts.test_cases.all():
            ts.test_cases.append(tc)
            db.session.commit()
            flash(f'Test case {tc.test_id_code} added to suite.', 'success')
    return redirect(url_for('suite_detail', suite_id=suite_id,
                            suite=request.form.get('suite_type', 'Functional')))


@app.route('/suites/<int:suite_id>/remove_test/<int:tc_id>', methods=['POST'])
def suite_remove_test(suite_id, tc_id):
    ts = TestSuite.query.get_or_404(suite_id)
    tc = TestCase.query.get_or_404(tc_id)
    if tc in ts.test_cases.all():
        ts.test_cases.remove(tc)
        db.session.commit()
        flash(f'Test case {tc.test_id_code} removed from suite.', 'success')
    return redirect(url_for('suite_detail', suite_id=suite_id,
                            suite=request.form.get('suite_type', 'Functional')))


@app.route('/suites/<int:suite_id>/import', methods=['GET', 'POST'])
def suite_import_tests(suite_id):
    """Import test cases from an Excel file into an existing suite template."""
    ts = TestSuite.query.get_or_404(suite_id)
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file (.xlsx).', 'danger')
            return redirect(url_for('suite_import_tests', suite_id=suite_id))

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        stats = import_excel(filepath, test_suite_id=suite_id)
        if 'error' in stats:
            flash(stats['error'], 'danger')
            return redirect(url_for('suite_import_tests', suite_id=suite_id))

        # Add newly imported test cases to this suite template
        new_ids = stats.get('test_ids', [])
        added = 0
        existing_tc_ids = {tc.id for tc in ts.test_cases.all()}
        for tc in TestCase.query.filter(TestCase.id.in_(new_ids)).all():
            if tc.id not in existing_tc_ids:
                ts.test_cases.append(tc)
                added += 1

        # Auto-create TestResults for vendors in projects using this suite
        new_results = 0
        for proj in ts.projects.all():
            for vendor in proj.vendors.all():
                existing_result_tc_ids = {tr.test_case_id for tr in vendor.results.all()}
                for tc_id in new_ids:
                    if tc_id not in existing_result_tc_ids:
                        db.session.add(TestResult(vendor_id=vendor.id, test_case_id=tc_id))
                        new_results += 1

        db.session.commit()

        total = stats['functional'] + stats['non_functional']
        flash(
            f'Imported {total} test cases ({stats["functional"]} functional, '
            f'{stats["non_functional"]} non-functional). '
            f'{added} added to "{ts.name}".'
            + (f' {new_results} vendor test results auto-created.' if new_results else ''),
            'success'
        )
        for w in stats.get('warnings', []):
            flash(w, 'warning')

        return redirect(url_for('suite_detail', suite_id=suite_id))

    return render_template('suite_import.html', ts=ts)




# ══════════════════════════════════════════════════════════════════════════════
#  VENDORS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/projects/<int:project_id>/vendors/new', methods=['GET', 'POST'])
def vendor_new(project_id):
    project = Project.query.get_or_404(project_id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        contact = request.form.get('contact', '').strip()
        notes = request.form.get('notes', '')
        eval_method = request.form.get('eval_method', 'Demo')
        if not name:
            # Return JSON if AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'ok': False, 'error': 'Vendor name is required.'}), 400
            flash('Vendor name is required.', 'danger')
            return redirect(url_for('vendor_new', project_id=project_id))
        v = Vendor(project_id=project_id, name=name, contact=contact, notes=notes, eval_method=eval_method)
        db.session.add(v)
        db.session.commit()

        # Auto-create test results for the project's test suite (or all tests if no suite)
        if project.test_suite_id:
            ts = TestSuite.query.get(project.test_suite_id)
            tests = ts.test_cases.order_by(TestCase.sort_order).all() if ts else TestCase.query.all()
        else:
            tests = TestCase.query.all()
        for tc in tests:
            tr = TestResult(vendor_id=v.id, test_case_id=tc.id, status='Not Started')
            db.session.add(tr)
        db.session.commit()

        # Return JSON if AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'ok': True, 'vendor_id': v.id, 'message': f'Vendor "{name}" added with {len(tests)} test results initialized.'})

        flash(f'Vendor "{name}" added with {len(tests)} test results initialized.', 'success')
        return redirect(url_for('project_detail', project_id=project_id))
    return render_template('vendor_form.html', project=project, vendor=None)


@app.route('/vendors/<int:vendor_id>/edit', methods=['GET', 'POST'])
def vendor_edit(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    project = vendor.project
    if request.method == 'POST':
        vendor.name = request.form.get('name', vendor.name).strip()
        vendor.contact = request.form.get('contact', '').strip()
        vendor.eval_method = request.form.get('eval_method', vendor.eval_method)
        vendor.notes = request.form.get('notes', '')
        db.session.commit()
        
        # Return JSON if AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'ok': True, 'vendor_id': vendor.id, 'message': f'Vendor "{vendor.name}" updated.'})
        
        flash(f'Vendor "{vendor.name}" updated.', 'success')
        return redirect(url_for('vendor_detail', vendor_id=vendor.id))
    return render_template('vendor_form.html', project=project, vendor=vendor)


@app.route('/api/vendors/<int:vendor_id>', methods=['GET'])
def api_vendor_get(vendor_id):
    """Return vendor data as JSON for edit modal."""
    vendor = Vendor.query.get_or_404(vendor_id)
    return jsonify({
        'id': vendor.id,
        'name': vendor.name,
        'contact': vendor.contact or '',
        'eval_method': vendor.eval_method,
        'notes': vendor.notes or ''
    })


@app.route('/vendors/<int:vendor_id>/delete', methods=['POST'])
def vendor_delete(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    project_id = vendor.project_id
    name = vendor.name

    # Delete all child data: evidence → audit → results
    for tr in vendor.results.all():
        Evidence.query.filter_by(test_result_id=tr.id).delete()
        AuditLog.query.filter_by(test_result_id=tr.id).delete()
        db.session.delete(tr)

    # Delete vendor-level comments and documents
    VendorComment.query.filter_by(vendor_id=vendor.id).delete()
    VendorDocument.query.filter_by(vendor_id=vendor.id).delete()

    _audit('deleted', 'Vendor', vendor.id)
    db.session.delete(vendor)
    db.session.commit()
    flash(f'Vendor "{name}" and all associated scores/evidence deleted.', 'success')
    return redirect(url_for('project_detail', project_id=project_id))


def _sync_vendor_results(vendor):
    """Ensure vendor has TestResult rows for every test in the project suite."""
    project = vendor.project
    if not project or not project.test_suite_id:
        return 0
    ts = TestSuite.query.get(project.test_suite_id)
    if not ts:
        return 0
    existing_tc_ids = {tr.test_case_id for tr in vendor.results.all()}
    suite_tc_ids = {tc.id for tc in ts.test_cases.all()}
    missing = suite_tc_ids - existing_tc_ids
    for tc_id in missing:
        db.session.add(TestResult(vendor_id=vendor.id, test_case_id=tc_id, status='Not Started'))
    if missing:
        db.session.commit()
    return len(missing)


@app.route('/vendors/<int:vendor_id>')
def vendor_detail(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    # Auto-sync: create TestResult rows for any new tests in the suite
    synced = _sync_vendor_results(vendor)
    if synced:
        flash(f'{synced} new test(s) synced from suite template.', 'info')
    suite = request.args.get('suite', 'Functional')
    tier = request.args.get('tier', '')
    priority = request.args.get('priority', '')
    area_filter = request.args.get('area', '')
    cat_filter = request.args.get('cat', '')
    subcat_filter = request.args.get('subcat', '')
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')

    q = db.session.query(TestResult, TestCase).join(TestCase).filter(
        TestResult.vendor_id == vendor_id,
        TestCase.suite_type == suite,
    )
    if tier:
        q = q.filter(TestCase.tier == tier)
    if priority:
        q = q.filter(TestCase.priority == priority)
    if subcat_filter:
        q = q.filter(TestCase.category_id == int(subcat_filter))
    elif cat_filter:
        cat_obj = Category.query.get(int(cat_filter))
        if cat_obj:
            sub_ids = [sc.id for sc in cat_obj.children.all()]
            q = q.filter(TestCase.category_id.in_([cat_obj.id] + sub_ids))
    elif area_filter:
        area_cat_ids = [c.id for c in Category.query.filter_by(area_id=int(area_filter)).all()]
        if area_cat_ids:
            q = q.filter(TestCase.category_id.in_(area_cat_ids))
    if status_filter:
        q = q.filter(TestResult.status == status_filter)
    if search:
        q = q.filter(db.or_(
            TestCase.test_id_code.ilike(f'%{search}%'),
            TestCase.capability.ilike(f'%{search}%'),
        ))

    results = q.order_by(TestCase.sort_order).all()

    # Pre-compute evidence & question counts to avoid N+1 queries
    result_ids = [tr.id for tr, tc in results]
    ev_counts = {}
    q_counts = {}
    open_q_counts = {}
    if result_ids:
        ev_rows = db.session.query(
            Evidence.test_result_id, db.func.count(Evidence.id)
        ).filter(Evidence.test_result_id.in_(result_ids)) \
         .group_by(Evidence.test_result_id).all()
        ev_counts = dict(ev_rows)

        q_rows = db.session.query(
            VendorQuestion.test_result_id, db.func.count(VendorQuestion.id)
        ).filter(VendorQuestion.test_result_id.in_(result_ids)) \
         .group_by(VendorQuestion.test_result_id).all()
        q_counts = dict(q_rows)

        open_q_rows = db.session.query(
            VendorQuestion.test_result_id, db.func.count(VendorQuestion.id)
        ).filter(
            VendorQuestion.test_result_id.in_(result_ids),
            VendorQuestion.status == 'Open'
        ).group_by(VendorQuestion.test_result_id).all()
        open_q_counts = dict(open_q_rows)

    # Scope areas to the project's suite template
    project = vendor.project
    ts = TestSuite.query.get(project.test_suite_id) if project.test_suite_id else None
    if ts:
        areas = Area.query.filter_by(test_suite_id=ts.id, suite_type=suite).order_by(Area.sort_order).all()
    else:
        areas = Area.query.filter_by(suite_type=suite).order_by(Area.sort_order).all()

    # Build cat_index, categories, subcategories for filter dropdowns
    categories = []
    subcategories = []
    cat_index = {}
    for ai, a in enumerate(areas, 1):
        cat_index[f'area_{a.id}'] = str(ai)
        ci = 0
        for c in Category.query.filter_by(area_id=a.id, parent_id=None).order_by(Category.sort_order).all():
            ci += 1
            categories.append(c)
            cat_index[c.id] = f'{ai}.{ci}'
            for si, sc in enumerate(c.children.order_by(Category.sort_order).all(), 1):
                subcategories.append(sc)
                cat_index[sc.id] = f'{ai}.{ci}.{si}'

    scoring_levels = ScoringLevel.query.order_by(ScoringLevel.sort_order).all()

    return render_template('vendor_detail.html', vendor=vendor, results=results,
                           areas=areas, categories=categories, subcategories=subcategories,
                           cat_index=cat_index, scoring_levels=scoring_levels,
                           suite=suite, tier=tier, priority=priority,
                           area_filter=area_filter, cat_filter=cat_filter,
                           subcat_filter=subcat_filter,
                           status_filter=status_filter, search=search,
                           ev_counts=ev_counts, q_counts=q_counts, open_q_counts=open_q_counts)


# ══════════════════════════════════════════════════════════════════════════════
#  ACTIVE EVALUATIONS  (single page listing all vendors ready to evaluate)
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/evaluations')
def active_evaluations():
    """Show all vendor evaluations across active projects."""
    vendors = db.session.query(Vendor).join(Project).filter(
        Project.status == 'Active'
    ).order_by(Project.name, Vendor.name).all()

    vendor_rows = []
    for v in vendors:
        _sync_vendor_results(v)  # ensure counts are up-to-date
        total = v.results.count()
        scored = v.results.filter(TestResult.score.isnot(None)).count() if total else 0
        total_ws = db.session.query(db.func.sum(TestResult.weighted_score)).filter(
            TestResult.vendor_id == v.id, TestResult.score.isnot(None)
        ).scalar() or 0
        max_possible = db.session.query(db.func.sum(TestCase.weight)).join(TestResult).filter(
            TestResult.vendor_id == v.id, TestResult.score.isnot(None)
        ).scalar() or 0
        max_score = 5
        avg_pct = round((total_ws / (max_possible * max_score) * 100) if max_possible else 0, 1)
        progress = round((scored / total * 100) if total else 0, 1)

        vendor_rows.append({
            'vendor': v,
            'project': v.project,
            'total': total,
            'scored': scored,
            'avg_pct': avg_pct,
            'progress': progress,
        })

    return render_template('evaluations.html', vendor_rows=vendor_rows)


# ══════════════════════════════════════════════════════════════════════════════
#  SCORING  (AJAX endpoints)
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/results/<int:result_id>/score', methods=['POST'])
def api_score_result(result_id):
    tr = TestResult.query.get_or_404(result_id)
    data = request.get_json() or request.form
    score_map = _scoring_map()

    old_level = tr.support_level
    old_score = tr.score

    support_level = data.get('support_level', tr.support_level)
    tr.support_level = support_level
    tr.score = score_map.get(support_level)
    tc = TestCase.query.get(tr.test_case_id)
    tr.weighted_score = (tr.score * tc.weight) if tr.score is not None else None
    tr.notes = data.get('notes', tr.notes)
    tr.status = data.get('status', tr.status)
    tr.pass_fail = data.get('pass_fail', tr.pass_fail)
    tr.block_reason = data.get('block_reason', tr.block_reason)

    _audit('scored', 'TestResult', tr.id, 'support_level', old_level, support_level, result_id=tr.id)
    db.session.commit()

    return jsonify({'ok': True, 'score': tr.score, 'weighted_score': tr.weighted_score})


@app.route('/api/results/bulk', methods=['POST'])
def api_bulk_update():
    data = request.get_json()
    ids = data.get('result_ids', [])
    updates = data.get('updates', {})
    score_map = _scoring_map()
    count = 0

    for rid in ids:
        tr = TestResult.query.get(rid)
        if not tr:
            continue
        old_vals = {
            'support_level': tr.support_level, 'score': tr.score,
            'status': tr.status, 'pass_fail': tr.pass_fail,
        }
        if updates.get('clear_score'):
            tr.support_level = None
            tr.score = None
            tr.weighted_score = None
        elif 'support_level' in updates:
            tr.support_level = updates['support_level']
            tr.score = score_map.get(tr.support_level)
            tc = TestCase.query.get(tr.test_case_id)
            tr.weighted_score = (tr.score * tc.weight) if tr.score is not None else None
        if 'status' in updates:
            tr.status = updates['status']
        if 'pass_fail' in updates:
            tr.pass_fail = updates['pass_fail']
        new_vals = {
            'support_level': tr.support_level, 'score': tr.score,
            'status': tr.status, 'pass_fail': tr.pass_fail,
        }
        _audit_changes('TestResult', tr.id, old_vals, new_vals, result_id=tr.id)
        count += 1

    db.session.commit()
    return jsonify({'ok': True, 'updated': count})


# ══════════════════════════════════════════════════════════════════════════════
#  EVIDENCE
# ══════════════════════════════════════════════════════════════════════════════
# ── Evidence JSON API (for modal) ────────────────────────────────────────────
@app.route('/api/results/<int:result_id>/evidence', methods=['GET'])
def api_evidence_list(result_id):
    """Return evidence items as JSON for the evidence modal."""
    tr = TestResult.query.get_or_404(result_id)
    tc = TestCase.query.get(tr.test_case_id)
    vendor = Vendor.query.get(tr.vendor_id)
    evidences = Evidence.query.filter_by(test_result_id=result_id).order_by(Evidence.uploaded_at.desc()).all()
    items = []
    for ev in evidences:
        item = {
            'id': ev.id,
            'type': ev.evidence_type,
            'uploaded_at': ev.uploaded_at.strftime('%Y-%m-%d %H:%M') if ev.uploaded_at else '',
        }
        if ev.evidence_type == 'file':
            item['filename'] = ev.filename
            item['url'] = url_for('serve_upload', vendor_id=vendor.id, filename=ev.filename)
        elif ev.evidence_type == 'link':
            item['url'] = ev.url or ''
        elif ev.evidence_type == 'text':
            item['text'] = ev.text_content or ''
        items.append(item)
    return jsonify({
        'test_id_code': tc.test_id_code,
        'capability': tc.capability or '',
        'evidence_required': tc.evidence_required or '',
        'vendor_name': vendor.name,
        'items': items,
    })


@app.route('/api/results/<int:result_id>/evidence', methods=['POST'])
def api_evidence_add(result_id):
    """Add evidence via AJAX (supports file upload as multipart or JSON for link/text)."""
    tr = TestResult.query.get_or_404(result_id)
    vendor = Vendor.query.get(tr.vendor_id)
    etype = request.form.get('evidence_type') or (request.get_json() or {}).get('evidence_type', '')

    if etype == 'file':
        file = request.files.get('file')
        if file and allowed_file(file.filename):
            fname = secure_filename(file.filename)
            vendor_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(vendor.id))
            os.makedirs(vendor_dir, exist_ok=True)
            fpath = os.path.join(vendor_dir, fname)
            file.save(fpath)
            ev = Evidence(test_result_id=result_id, evidence_type='file',
                          filename=fname, filepath=fpath,
                          uploaded_by=current_user.email if current_user.is_authenticated else 'system')
            db.session.add(ev)
            db.session.commit()
            return jsonify({'ok': True, 'id': ev.id})
        return jsonify({'ok': False, 'error': 'Invalid file type.'}), 400
    elif etype == 'link':
        data = request.get_json() if request.is_json else {'url': request.form.get('url', '')}
        url_val = data.get('url', '').strip()
        if url_val:
            ev = Evidence(test_result_id=result_id, evidence_type='link', url=url_val,
                          uploaded_by=current_user.email if current_user.is_authenticated else 'system')
            db.session.add(ev)
            db.session.commit()
            return jsonify({'ok': True, 'id': ev.id})
        return jsonify({'ok': False, 'error': 'URL is required.'}), 400
    elif etype == 'text':
        data = request.get_json() if request.is_json else {'text_content': request.form.get('text_content', '')}
        text = data.get('text_content', '').strip()
        if text:
            ev = Evidence(test_result_id=result_id, evidence_type='text', text_content=text,
                          uploaded_by=current_user.email if current_user.is_authenticated else 'system')
            db.session.add(ev)
            db.session.commit()
            return jsonify({'ok': True, 'id': ev.id})
        return jsonify({'ok': False, 'error': 'Text is required.'}), 400

    return jsonify({'ok': False, 'error': 'Unknown evidence type.'}), 400


@app.route('/api/evidence/<int:evidence_id>', methods=['DELETE'])
def api_evidence_delete(evidence_id):
    """Delete a single evidence item."""
    ev = Evidence.query.get_or_404(evidence_id)
    db.session.delete(ev)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/vendors/<int:vendor_id>/evidence', methods=['GET'])
def api_vendor_evidence(vendor_id):
    """Return ALL evidence items across all test results for a vendor."""
    vendor = Vendor.query.get_or_404(vendor_id)
    rows = db.session.query(Evidence, TestResult, TestCase).join(
        TestResult, Evidence.test_result_id == TestResult.id
    ).join(
        TestCase, TestResult.test_case_id == TestCase.id
    ).filter(
        TestResult.vendor_id == vendor_id
    ).order_by(Evidence.uploaded_at.desc()).all()

    items = []
    for ev, tr, tc in rows:
        item = {
            'id': ev.id,
            'type': ev.evidence_type,
            'test_id_code': tc.test_id_code,
            'capability': tc.capability or '',
            'test_result_id': tr.id,
            'uploaded_at': ev.uploaded_at.strftime('%Y-%m-%d %H:%M') if ev.uploaded_at else '',
            'uploaded_by': ev.uploaded_by or '',
        }
        if ev.evidence_type == 'file':
            item['filename'] = ev.filename
            item['url'] = url_for('serve_upload', vendor_id=vendor.id, filename=ev.filename)
        elif ev.evidence_type == 'link':
            item['url'] = ev.url or ''
        elif ev.evidence_type == 'text':
            item['text'] = ev.text_content or ''
        items.append(item)
    return jsonify({'items': items})


@app.route('/uploads/<int:vendor_id>/<filename>')
def serve_upload(vendor_id, filename):
    vendor_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(vendor_id))
    return send_from_directory(vendor_dir, filename)


# ══════════════════════════════════════════════════════════════════════════════
#  VENDOR COMMENTS & DOCUMENTS  (meeting notes, uploaded files per vendor)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/vendors/<int:vendor_id>/comments', methods=['GET'])
def api_vendor_comments(vendor_id):
    """Return all comments for a vendor as JSON."""
    vendor = Vendor.query.get_or_404(vendor_id)
    comments = VendorComment.query.filter_by(vendor_id=vendor_id) \
        .order_by(VendorComment.created_at.desc()).all()
    items = []
    for c in comments:
        items.append({
            'id': c.id,
            'title': c.title or '',
            'body': c.body,
            'created_at': c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
            'updated_at': c.updated_at.strftime('%Y-%m-%d %H:%M') if c.updated_at else '',
            'created_by': c.created_by or '',
        })
    return jsonify({'ok': True, 'items': items})


@app.route('/api/vendors/<int:vendor_id>/comments', methods=['POST'])
def api_vendor_comment_add(vendor_id):
    """Add a new comment / meeting note for a vendor."""
    vendor = Vendor.query.get_or_404(vendor_id)
    data = request.get_json() if request.is_json else {}
    title = (data.get('title') or request.form.get('title', '')).strip()
    body = (data.get('body') or request.form.get('body', '')).strip()
    if not body:
        return jsonify({'ok': False, 'error': 'Comment body is required.'}), 400
    comment = VendorComment(vendor_id=vendor_id, title=title, body=body,
                             created_by=current_user.email if current_user.is_authenticated else 'system')
    db.session.add(comment)
    db.session.commit()
    _audit('created', 'VendorComment', comment.id)
    return jsonify({'ok': True, 'id': comment.id,
                    'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M')})


@app.route('/api/vendor-comments/<int:comment_id>', methods=['PUT'])
def api_vendor_comment_edit(comment_id):
    """Edit an existing vendor comment."""
    comment = VendorComment.query.get_or_404(comment_id)
    data = request.get_json() if request.is_json else {}
    title = data.get('title', comment.title or '').strip()
    body = data.get('body', comment.body).strip()
    if not body:
        return jsonify({'ok': False, 'error': 'Comment body is required.'}), 400
    comment.title = title
    comment.body = body
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/vendor-comments/<int:comment_id>', methods=['DELETE'])
def api_vendor_comment_delete(comment_id):
    """Delete a vendor comment and any inline images it references."""
    comment = VendorComment.query.get_or_404(comment_id)
    # Extract image filenames from the comment HTML body
    if comment.body:
        img_srcs = re.findall(r'<img[^>]+src=["\'](?:/uploads/\d+/)?([^"\'/]+)["\']', comment.body)
        if img_srcs:
            vendor_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(comment.vendor_id))
            for fname in img_srcs:
                # Delete matching VendorDocument record
                doc = VendorDocument.query.filter_by(
                    vendor_id=comment.vendor_id, filename=fname
                ).first()
                if doc:
                    db.session.delete(doc)
                # Delete the physical file
                fpath = os.path.join(vendor_dir, fname)
                if os.path.isfile(fpath):
                    try:
                        os.remove(fpath)
                    except OSError:
                        pass
    db.session.delete(comment)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/vendors/<int:vendor_id>/upload-image', methods=['POST'])
def api_vendor_upload_image(vendor_id):
    """Upload an image for inline embedding in comments. Returns the image URL."""
    vendor = Vendor.query.get_or_404(vendor_id)
    file = request.files.get('file')
    if not file:
        return jsonify({'ok': False, 'error': 'No file provided.'}), 400
    if not allowed_file(file.filename):
        return jsonify({'ok': False, 'error': 'File type not allowed.'}), 400
    fname = secure_filename(file.filename)
    # Ensure unique filename with timestamp
    import time
    base, ext = os.path.splitext(fname)
    fname = f"{base}_{int(time.time())}{ext}"
    vendor_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(vendor.id))
    os.makedirs(vendor_dir, exist_ok=True)
    fpath = os.path.join(vendor_dir, fname)
    file.save(fpath)
    # Also create a VendorDocument so the image appears in the gallery
    doc = VendorDocument(
        vendor_id=vendor.id,
        doc_type='file',
        filename=fname,
        filepath=fpath,
        description='',
        uploaded_by=current_user.email if current_user.is_authenticated else 'system',
    )
    db.session.add(doc)
    db.session.commit()
    img_url = url_for('serve_upload', vendor_id=vendor.id, filename=fname)
    return jsonify({'ok': True, 'url': img_url, 'filename': fname})


@app.route('/api/vendors/<int:vendor_id>/documents', methods=['GET'])
def api_vendor_documents(vendor_id):
    """Return all documents for a vendor as JSON."""
    vendor = Vendor.query.get_or_404(vendor_id)
    docs = VendorDocument.query.filter_by(vendor_id=vendor_id) \
        .order_by(VendorDocument.uploaded_at.desc()).all()
    items = []
    for d in docs:
        item = {
            'id': d.id,
            'doc_type': d.doc_type,
            'description': d.description or '',
            'uploaded_at': d.uploaded_at.strftime('%Y-%m-%d %H:%M') if d.uploaded_at else '',
            'uploaded_by': d.uploaded_by or '',
        }
        if d.doc_type == 'file':
            item['filename'] = d.filename
            item['url'] = url_for('serve_upload', vendor_id=vendor_id, filename=d.filename)
        elif d.doc_type == 'link':
            item['url'] = d.url or ''
        items.append(item)
    return jsonify({'ok': True, 'items': items})


@app.route('/api/vendors/<int:vendor_id>/documents', methods=['POST'])
def api_vendor_document_add(vendor_id):
    """Upload a file or add a link as a vendor-level document."""
    vendor = Vendor.query.get_or_404(vendor_id)
    if request.is_json:
        data = request.get_json()
        dtype = data.get('doc_type', 'file')
        description = data.get('description', '')
    else:
        dtype = request.form.get('doc_type', 'file')
        description = request.form.get('description', '')

    if dtype == 'file':
        file = request.files.get('file')
        if file and allowed_file(file.filename):
            fname = secure_filename(file.filename)
            vendor_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(vendor.id))
            os.makedirs(vendor_dir, exist_ok=True)
            fpath = os.path.join(vendor_dir, fname)
            file.save(fpath)
            doc = VendorDocument(vendor_id=vendor_id, doc_type='file',
                                filename=fname, filepath=fpath, description=description.strip(),
                                uploaded_by=current_user.email if current_user.is_authenticated else 'system')
            db.session.add(doc)
            db.session.commit()
            _audit('created', 'VendorDocument', doc.id)
            return jsonify({'ok': True, 'id': doc.id})
        return jsonify({'ok': False, 'error': 'Invalid or missing file.'}), 400
    elif dtype == 'link':
        data = request.get_json() if request.is_json else {'url': request.form.get('url', '')}
        url_val = (data.get('url') or '').strip()
        if url_val:
            doc = VendorDocument(vendor_id=vendor_id, doc_type='link',
                                url=url_val, description=description.strip(),
                                uploaded_by=current_user.email if current_user.is_authenticated else 'system')
            db.session.add(doc)
            db.session.commit()
            _audit('created', 'VendorDocument', doc.id)
            return jsonify({'ok': True, 'id': doc.id})
        return jsonify({'ok': False, 'error': 'URL is required.'}), 400

    return jsonify({'ok': False, 'error': 'Unknown document type.'}), 400


@app.route('/api/vendor-documents/<int:doc_id>', methods=['DELETE'])
def api_vendor_document_delete(doc_id):
    """Delete a vendor document."""
    doc = VendorDocument.query.get_or_404(doc_id)
    # Try to remove the physical file if it's a file type
    if doc.doc_type == 'file' and doc.filepath:
        try:
            os.remove(doc.filepath)
        except OSError:
            pass
    db.session.delete(doc)
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
#  REPORTING & COMPARISON
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/projects/<int:project_id>/scorecard')
def project_scorecard(project_id):
    project = Project.query.get_or_404(project_id)
    vendors = project.vendors.all()

    # Build scorecard data
    scorecard = []
    for vendor in vendors:
        vendor_data = {
            'vendor': vendor,
            'total_weighted': 0,
            'total_weight': 0,
            'scored_count': 0,
            'total_count': 0,
            'mandatory_fails': [],
            'categories': {},
        }

        results = db.session.query(TestResult, TestCase).join(TestCase).filter(
            TestResult.vendor_id == vendor.id
        ).all()

        for tr, tc in results:
            vendor_data['total_count'] += 1

            cat = Category.query.get(tc.category_id)
            # Walk up to the Area name for grouping
            root_cat = cat
            while root_cat.parent_id:
                root_cat = Category.query.get(root_cat.parent_id)
            area = Area.query.get(root_cat.area_id) if root_cat.area_id else None
            group_name = area.name if area else root_cat.name

            if group_name not in vendor_data['categories']:
                vendor_data['categories'][group_name] = {
                    'suite_type': tc.suite_type,
                    'total_weight': 0,
                    'total_weighted': 0,
                    'scored': 0,
                    'total': 0,
                    'mandatory_fails': [],
                }

            cat_data = vendor_data['categories'][group_name]
            cat_data['total'] += 1
            cat_data['total_weight'] += tc.weight

            if tr.score is not None:
                ws = tr.weighted_score or 0
                vendor_data['total_weighted'] += ws
                vendor_data['total_weight'] += tc.weight
                vendor_data['scored_count'] += 1
                cat_data['total_weighted'] += ws
                cat_data['scored'] += 1

            if tc.is_mandatory and (tr.score is not None and tr.score == 0):
                vendor_data['mandatory_fails'].append(tc)
                cat_data['mandatory_fails'].append(tc)

        scorecard.append(vendor_data)

    return render_template('scorecard.html', project=project, scorecard=scorecard)


@app.route('/projects/<int:project_id>/compare')
def project_compare(project_id):
    project = Project.query.get_or_404(project_id)
    vendors = project.vendors.all()
    suite = request.args.get('suite', 'Functional')
    tier = request.args.get('tier', '')
    group_by = request.args.get('group_by', 'area')  # 'area' or 'category'

    score_map = _scoring_map()
    max_score_val = max((v for v in score_map.values() if v is not None), default=5)

    def _vendor_stats(vendor, cat_ids, tier):
        """Compute vendor comparison stats for a set of category ids."""
        results = db.session.query(TestResult, TestCase).join(TestCase).filter(
            TestResult.vendor_id == vendor.id,
            TestCase.category_id.in_(cat_ids),
        )
        if tier:
            results = results.filter(TestCase.tier == tier)
        results = results.all()
        total_weight = sum(tc.weight for tr, tc in results if tr.score is not None)
        total_ws = sum((tr.weighted_score or 0) for tr, tc in results if tr.score is not None)
        scored = sum(1 for tr, _ in results if tr.score is not None)
        avg = (total_ws / (total_weight * max_score_val) * 100) if total_weight else 0
        return {
            'avg_pct': round(avg, 1),
            'scored': scored,
            'total': len(results),
            'total_ws': round(total_ws, 1),
            'total_weight': total_weight,
        }

    comparison = []
    # Scope areas to the project's suite template
    ts = TestSuite.query.get(project.test_suite_id) if project.test_suite_id else None
    _area_filter = {'suite_type': suite}
    if ts:
        _area_filter['test_suite_id'] = ts.id

    if group_by == 'category':
        # Group by top-level Category within each Area for the selected suite
        areas = Area.query.filter_by(**_area_filter).order_by(Area.sort_order).all()
        for area in areas:
            area_cats = Category.query.filter_by(area_id=area.id, parent_id=None).order_by(Category.sort_order).all()
            for cat in area_cats:
                cat_ids = [cat.id] + [sub.id for sub in cat.children.all()]
                row = {'area': area.name, 'label': cat.name, 'vendors': {}}
                for vendor in vendors:
                    row['vendors'][vendor.name] = _vendor_stats(vendor, cat_ids, tier)
                comparison.append(row)
        # Also pick up categories not assigned to any area
        orphan_cats = Category.query.filter_by(suite_type=suite, parent_id=None, area_id=None).order_by(Category.sort_order).all()
        for cat in orphan_cats:
            cat_ids = [cat.id] + [sub.id for sub in cat.children.all()]
            row = {'area': '(No Area)', 'label': cat.name, 'vendors': {}}
            for vendor in vendors:
                row['vendors'][vendor.name] = _vendor_stats(vendor, cat_ids, tier)
            comparison.append(row)

        # Pre-compute area_span / area_first for rowspan grouping in the template
        i = 0
        while i < len(comparison):
            area_name = comparison[i]['area']
            span = sum(1 for r in comparison[i:] if r['area'] == area_name)
            comparison[i]['area_first'] = True
            comparison[i]['area_span'] = span
            for j in range(i + 1, i + span):
                comparison[j]['area_first'] = False
                comparison[j]['area_span'] = 0
            i += span
    else:
        # Group by Area (default)
        areas = Area.query.filter_by(**_area_filter).order_by(Area.sort_order).all()
        for area in areas:
            cat_ids = []
            for cat in area.categories.all():
                cat_ids.append(cat.id)
                for sub in cat.children.all():
                    cat_ids.append(sub.id)
            if not cat_ids:
                continue
            row = {'label': area.name, 'vendors': {}}
            for vendor in vendors:
                row['vendors'][vendor.name] = _vendor_stats(vendor, cat_ids, tier)
            comparison.append(row)

    return render_template('compare.html', project=project, vendors=vendors,
                           comparison=comparison, suite=suite, tier=tier,
                           group_by=group_by)


# ══════════════════════════════════════════════════════════════════════════════
#  SCORING SCALE MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/settings/scoring', methods=['GET', 'POST'])
def scoring_settings():
    if request.method == 'POST':
        # Clear and re-create
        ScoringLevel.query.delete()
        labels = request.form.getlist('label')
        values = request.form.getlist('value')
        for i, (label, val) in enumerate(zip(labels, values)):
            label = label.strip()
            if not label:
                continue
            try:
                v = int(val) if val.strip() != '' else None
            except ValueError:
                v = None
            db.session.add(ScoringLevel(label=label, value=v, sort_order=i))
        db.session.commit()
        flash('Scoring scale updated.', 'success')
        return redirect(url_for('scoring_settings'))

    levels = ScoringLevel.query.order_by(ScoringLevel.sort_order).all()
    return render_template('scoring_settings.html', levels=levels)


@app.route('/api/suites/<int:suite_id>/areas', methods=['POST'])
def api_area_create(suite_id):
    """Create a new area via JSON."""
    ts = TestSuite.query.get_or_404(suite_id)
    data = request.get_json()
    name = data.get('name', '').strip()
    suite_type = data.get('suite_type', 'Functional')
    if not name:
        return jsonify({'ok': False, 'error': 'Name is required'}), 400
    max_sort = db.session.query(db.func.max(Area.sort_order)).filter_by(
        test_suite_id=suite_id, suite_type=suite_type).scalar() or 0
    area = Area(name=name, suite_type=suite_type, test_suite_id=suite_id, sort_order=max_sort + 1)
    db.session.add(area)
    _audit('created', 'Area', 0)
    db.session.commit()
    return jsonify({'ok': True, 'id': area.id, 'name': area.name})


@app.route('/api/suites/<int:suite_id>/categories', methods=['POST'])
def api_category_create(suite_id):
    """Create a new category/subcategory via JSON."""
    ts = TestSuite.query.get_or_404(suite_id)
    data = request.get_json()
    name = data.get('name', '').strip()
    suite_type = data.get('suite_type', 'Functional')
    if not name:
        return jsonify({'ok': False, 'error': 'Name is required'}), 400
    area_id = data.get('area_id')
    parent_id = data.get('parent_id')
    # Prevent orphan categories: must have area or parent
    if not area_id and not parent_id:
        return jsonify({'ok': False, 'error': 'Area is required. Please select an area.'}), 400
    # If parent_id is given but no area_id, inherit area from the parent
    if parent_id and not area_id:
        parent_cat = Category.query.get(int(parent_id))
        if parent_cat:
            # Walk up to find the effective area
            pc = parent_cat
            while pc and not pc.area_id:
                pc = Category.query.get(pc.parent_id) if pc.parent_id else None
            if pc and pc.area_id:
                area_id = pc.area_id
    weight_mult = float(data.get('weight_multiplier', 1.0))
    max_sort = db.session.query(db.func.max(Category.sort_order)).filter_by(suite_type=suite_type).scalar() or 0
    cat = Category(
        name=name, suite_type=suite_type,
        area_id=int(area_id) if area_id else None,
        parent_id=int(parent_id) if parent_id else None,
        weight_multiplier=weight_mult, sort_order=max_sort + 1,
    )
    db.session.add(cat)
    _audit('created', 'Category', 0)
    db.session.commit()
    return jsonify({'ok': True, 'id': cat.id, 'name': cat.name})


@app.route('/api/areas/<int:area_id>', methods=['GET', 'PUT'])
def api_area(area_id):
    """Get or update an area via JSON."""
    area = Area.query.get_or_404(area_id)
    if request.method == 'GET':
        return jsonify({'id': area.id, 'name': area.name, 'suite_type': area.suite_type})
    data = request.get_json()
    old_name = area.name
    area.name = data.get('name', area.name).strip()
    _audit('updated', 'Area', area.id, 'name', old_name, area.name)
    db.session.commit()
    return jsonify({'ok': True, 'id': area.id, 'name': area.name})


@app.route('/api/categories/<int:cat_id>', methods=['GET', 'PUT'])
def api_category(cat_id):
    """Get or update a category/subcategory via JSON."""
    cat = Category.query.get_or_404(cat_id)

    # Resolve the effective area – subcategories may have area_id=None
    # so we walk up the parent chain to find the area.
    def _resolve_area(c):
        visited = set()
        while c:
            if c.id in visited:
                break
            visited.add(c.id)
            if c.area_id:
                return Area.query.get(c.area_id)
            c = Category.query.get(c.parent_id) if c.parent_id else None
        return None

    if request.method == 'GET':
        area_obj = _resolve_area(cat)
        ts = TestSuite.query.get(area_obj.test_suite_id) if area_obj and area_obj.test_suite_id else None
        suite_id = ts.id if ts else None
        all_areas = Area.query.filter_by(test_suite_id=suite_id, suite_type=cat.suite_type).order_by(Area.sort_order).all() if suite_id else []
        # Build numbering index: area_num[area_id] and cat_num[cat_id]
        area_num = {}
        cat_num = {}
        for ai, a in enumerate(all_areas, 1):
            area_num[a.id] = ai
            ci = 0
            for c in Category.query.filter_by(area_id=a.id, parent_id=None).order_by(Category.sort_order).all():
                ci += 1
                cat_num[c.id] = f'{ai}.{ci}'
        areas = [{'id': a.id, 'name': f"{area_num[a.id]}. {a.name}"} for a in all_areas]
        area_ids = [a['id'] for a in areas]
        parents = [{'id': p.id, 'area_id': p.area_id,
                    'name': f"{cat_num.get(p.id, '')}. {p.name}" if cat_num.get(p.id) else p.name}
                   for p in Category.query.filter(
                       Category.area_id.in_(area_ids),
                       Category.parent_id.is_(None),
                       Category.id != cat.id,
                   ).order_by(Category.sort_order).all()]
        effective_area_id = area_obj.id if area_obj else cat.area_id
        return jsonify({
            'id': cat.id, 'name': cat.name, 'suite_type': cat.suite_type,
            'area_id': effective_area_id, 'parent_id': cat.parent_id,
            'weight_multiplier': cat.weight_multiplier,
            'areas': areas, 'parents': parents,
        })
    # ── PUT ──
    data = request.get_json()
    old_vals = {
        'name': cat.name, 'weight_multiplier': cat.weight_multiplier,
        'area_id': cat.area_id, 'parent_id': cat.parent_id,
    }
    cat.name = data.get('name', cat.name).strip()
    cat.weight_multiplier = float(data.get('weight_multiplier', cat.weight_multiplier))
    # Only change area_id / parent_id when the client sends a real value;
    # an empty string or missing key means "keep current".
    area_id = data.get('area_id')
    if area_id not in (None, '', False):
        cat.area_id = int(area_id)
    parent_id = data.get('parent_id')
    if parent_id not in (None, '', False):
        cat.parent_id = int(parent_id)
    elif 'parent_id' in data and not parent_id:
        # Explicitly sent empty/null → promote to top-level only if
        # the category currently IS a top-level category (i.e. don't
        # orphan subcategories by accident).
        if cat.parent_id is None:
            cat.parent_id = None  # no-op, already top-level
        # else: keep existing parent_id (don't orphan)
    new_vals = {
        'name': cat.name, 'weight_multiplier': cat.weight_multiplier,
        'area_id': cat.area_id, 'parent_id': cat.parent_id,
    }
    _audit_changes('Category', cat.id, old_vals, new_vals)
    db.session.commit()
    return jsonify({'ok': True, 'id': cat.id, 'name': cat.name})


@app.route('/api/areas/<int:area_id>/move', methods=['POST'])
def api_area_move(area_id):
    """Move an area up or down within its suite."""
    area = Area.query.get_or_404(area_id)
    direction = request.get_json().get('direction')  # 'up' or 'down'
    siblings = Area.query.filter_by(
        test_suite_id=area.test_suite_id, suite_type=area.suite_type
    ).order_by(Area.sort_order, Area.id).all()
    # Normalise sort_order so every sibling has a unique sequential value
    for i, s in enumerate(siblings):
        s.sort_order = i
    db.session.flush()
    idx = next((i for i, a in enumerate(siblings) if a.id == area_id), None)
    if idx is None:
        return jsonify({'ok': False}), 404
    swap_idx = idx - 1 if direction == 'up' else idx + 1
    if swap_idx < 0 or swap_idx >= len(siblings):
        return jsonify({'ok': False, 'error': 'Already at boundary'}), 400
    siblings[idx].sort_order, siblings[swap_idx].sort_order = siblings[swap_idx].sort_order, siblings[idx].sort_order
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/categories/<int:cat_id>/move', methods=['POST'])
def api_category_move(cat_id):
    """Move a category or subcategory up or down among its siblings."""
    cat = Category.query.get_or_404(cat_id)
    direction = request.get_json().get('direction')
    siblings = Category.query.filter_by(
        area_id=cat.area_id, parent_id=cat.parent_id, suite_type=cat.suite_type
    ).order_by(Category.sort_order, Category.id).all()
    # Normalise sort_order so every sibling has a unique sequential value
    for i, s in enumerate(siblings):
        s.sort_order = i
    db.session.flush()
    idx = next((i for i, c in enumerate(siblings) if c.id == cat_id), None)
    if idx is None:
        return jsonify({'ok': False}), 404
    swap_idx = idx - 1 if direction == 'up' else idx + 1
    if swap_idx < 0 or swap_idx >= len(siblings):
        return jsonify({'ok': False, 'error': 'Already at boundary'}), 400
    siblings[idx].sort_order, siblings[swap_idx].sort_order = siblings[swap_idx].sort_order, siblings[idx].sort_order
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
#  VENDOR QUESTIONS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/projects/<int:project_id>/questions')
def project_questions(project_id):
    """Aggregation page: all questions for a project, filterable by vendor and area."""
    project = Project.query.get_or_404(project_id)
    vendors = project.vendors.order_by(Vendor.name).all()
    ts = TestSuite.query.get(project.test_suite_id) if project.test_suite_id else None
    if ts:
        areas = Area.query.filter_by(test_suite_id=ts.id).order_by(Area.sort_order).all()
    else:
        areas = Area.query.order_by(Area.sort_order).all()

    # Filters
    vendor_filter = request.args.get('vendor', '', type=str)
    area_filter = request.args.get('area', '', type=str)
    status_filter = request.args.get('status', '', type=str)

    q = VendorQuestion.query.filter_by(project_id=project_id)
    if vendor_filter:
        q = q.filter_by(vendor_id=int(vendor_filter))
    if area_filter:
        q = q.filter_by(area_id=int(area_filter))
    if status_filter:
        q = q.filter_by(status=status_filter)

    questions = q.order_by(VendorQuestion.created_at.desc()).all()

    # Enrich each question with traceability info
    question_rows = []
    for qn in questions:
        row = {
            'question': qn,
            'vendor_name': qn.vendor.name if qn.vendor else 'All Vendors',
            'area_name': qn.area.name if qn.area else '—',
            'category_name': qn.category.name if qn.category else '—',
            'test_id_code': None,
            'capability': None,
        }
        if qn.test_result_id:
            tr = TestResult.query.get(qn.test_result_id)
            if tr:
                tc = TestCase.query.get(tr.test_case_id)
                if tc:
                    row['test_id_code'] = tc.test_id_code
                    row['capability'] = tc.capability
        question_rows.append(row)

    return render_template('questions.html', project=project, vendors=vendors,
                           areas=areas, question_rows=question_rows,
                           vendor_filter=vendor_filter, area_filter=area_filter,
                           status_filter=status_filter)


@app.route('/api/projects/<int:project_id>/questions', methods=['POST'])
def api_question_create(project_id):
    """Create a new question for vendor."""
    project = Project.query.get_or_404(project_id)
    data = request.get_json()
    text = data.get('question_text', '').strip()
    if not text:
        return jsonify({'ok': False, 'error': 'Question text is required'}), 400

    vendor_id = data.get('vendor_id')
    test_result_id = data.get('test_result_id')
    area_id = data.get('area_id')
    category_id = data.get('category_id')

    # Auto-resolve area/category from test_result if provided
    if test_result_id and (not area_id or not category_id):
        tr = TestResult.query.get(int(test_result_id))
        if tr:
            tc = TestCase.query.get(tr.test_case_id)
            if tc:
                cat = Category.query.get(tc.category_id)
                if cat:
                    if not category_id:
                        category_id = cat.id
                    # Walk up to find area
                    if not area_id:
                        c = cat
                        while c:
                            if c.area_id:
                                area_id = c.area_id
                                break
                            c = Category.query.get(c.parent_id) if c.parent_id else None

    qn = VendorQuestion(
        project_id=project_id,
        vendor_id=int(vendor_id) if vendor_id else None,
        test_result_id=int(test_result_id) if test_result_id else None,
        area_id=int(area_id) if area_id else None,
        category_id=int(category_id) if category_id else None,
        question_text=text,
    )
    db.session.add(qn)
    _audit('created', 'VendorQuestion', 0)
    db.session.commit()
    return jsonify({'ok': True, 'id': qn.id})


@app.route('/api/questions/<int:question_id>', methods=['PUT'])
def api_question_update(question_id):
    """Update a question (edit text or record vendor response)."""
    qn = VendorQuestion.query.get_or_404(question_id)
    data = request.get_json()

    if 'question_text' in data:
        qn.question_text = data['question_text'].strip()
    if 'vendor_response' in data:
        response = data['vendor_response']
        # Handle null or empty response (delete answer)
        if response is None or (isinstance(response, str) and not response.strip()):
            qn.vendor_response = None
            qn.responded_at = None
            qn.status = 'Open'
        else:
            qn.vendor_response = response.strip()
            qn.responded_at = datetime.now(timezone.utc)
            qn.status = 'Answered'
    if 'status' in data:
        qn.status = data['status']

    db.session.commit()
    return jsonify({'ok': True, 'id': qn.id, 'status': qn.status})


@app.route('/api/questions/<int:question_id>', methods=['DELETE'])
def api_question_delete(question_id):
    """Delete a question."""
    qn = VendorQuestion.query.get_or_404(question_id)
    db.session.delete(qn)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/results/<int:result_id>/questions', methods=['GET'])
def api_result_questions(result_id):
    """Get all questions for a specific test result."""
    tr = TestResult.query.get_or_404(result_id)
    questions = VendorQuestion.query.filter_by(test_result_id=result_id).order_by(VendorQuestion.created_at).all()
    return jsonify({
        'items': [{
            'id': q.id,
            'question_text': q.question_text,
            'vendor_response': q.vendor_response or '',
            'status': q.status,
            'created_at': q.created_at.strftime('%Y-%m-%d %H:%M') if q.created_at else '',
            'responded_at': q.responded_at.strftime('%Y-%m-%d %H:%M') if q.responded_at else '',
        } for q in questions],
    })


@app.route('/api/scoring/reorder', methods=['POST'])
def api_scoring_reorder():
    """Accept a JSON list of level IDs in desired order and update sort_order."""
    ids = request.get_json()  # [id1, id2, ...]
    if not ids or not isinstance(ids, list):
        return jsonify({'ok': False, 'error': 'Expected a JSON list of IDs'}), 400
    for idx, level_id in enumerate(ids):
        sl = ScoringLevel.query.get(level_id)
        if sl:
            sl.sort_order = idx
    db.session.commit()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
#  AUDIT LOG
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/audit')
def audit_log():
    page = request.args.get('page', 1, type=int)
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).paginate(page=page, per_page=50)
    return render_template('audit_log.html', logs=logs)


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/export/template')
def download_import_template():
    """Download a blank Excel template with the correct column headers for importing test cases."""
    import openpyxl
    from io import BytesIO
    from flask import send_file
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill_f = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_fill_nf = PatternFill(start_color='FD7E14', end_color='FD7E14', fill_type='solid')
    lookup_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')

    headers = ['Test ID', 'Tier', 'Area', 'Category', 'Subcategory',
               'Capability / Feature', 'Test Scenario', 'Pass Criteria',
               'Evidence Required', 'Test Method', 'Priority', 'Weight']
    widths = [12, 10, 20, 25, 25, 45, 40, 35, 35, 15, 12, 10]

    # ── Build Lookups sheet ─────────────────────────────────────────
    ws_lk = wb.create_sheet(title='Lookups')
    lk_headers = ['Area', 'Category', 'Subcategory']
    ws_lk.append(lk_headers)
    for ci in range(1, 4):
        cell = ws_lk.cell(row=1, column=ci)
        cell.font = header_font
        cell.fill = lookup_fill
        cell.alignment = Alignment(horizontal='center')

    all_areas = Area.query.order_by(Area.suite_type, Area.sort_order).all()
    lookup_rows = []
    for area in all_areas:
        for cat in area.categories.filter_by(parent_id=None).order_by(Category.sort_order).all():
            subs = cat.children.order_by(Category.sort_order).all()
            if subs:
                for sub in subs:
                    lookup_rows.append((area.name, cat.name, sub.name))
            else:
                lookup_rows.append((area.name, cat.name, ''))
    orphan_cats = Category.query.filter_by(area_id=None, parent_id=None).order_by(Category.sort_order).all()
    for cat in orphan_cats:
        subs = cat.children.order_by(Category.sort_order).all()
        if subs:
            for sub in subs:
                lookup_rows.append(('', cat.name, sub.name))
        else:
            lookup_rows.append(('', cat.name, ''))
    for row in lookup_rows:
        ws_lk.append(list(row))
    for col in ws_lk.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws_lk.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    area_names = sorted(set(r[0] for r in lookup_rows if r[0]))
    cat_names = sorted(set(r[1] for r in lookup_rows if r[1]))
    sub_names = sorted(set(r[2] for r in lookup_rows if r[2]))

    def _dv_list(values, max_len=255):
        formula = '"' + ','.join(values) + '"'
        if len(formula) <= max_len:
            return DataValidation(type='list', formula1=formula, allow_blank=True,
                                 showErrorMessage=True, errorTitle='Invalid value',
                                 error='Please select a value from the list.')
        return None

    # ── Build test suite sheets with example row + validation ──────
    hints = ['F-001', 'Core', area_names[0] if area_names else 'My Area',
             cat_names[0] if cat_names else 'My Category',
             sub_names[0] if sub_names else '',
             'Support for OCPP 2.0.1', 'Demonstrate in dashboard...',
             'Profile visible in UI', 'Screenshot of config screen',
             'Demo', 'Must', '1.0']

    for sheet_name, fill in [('Functional Test Suite', header_fill_f), ('Non-Functional Test Suite', header_fill_nf)]:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]
        # Example row
        prefix = 'F' if 'Functional' == sheet_name.split()[0] else 'NF'
        example = list(hints)
        example[0] = f'{prefix}-001'
        ws.append(example)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=2, column=col_idx).font = Font(italic=True, color='999999')

        max_row = 500
        # Data validations
        tier_dv = DataValidation(type='list', formula1='"Core,Extended"', allow_blank=True)
        tier_dv.add(f'B2:B{max_row}')
        ws.add_data_validation(tier_dv)

        priority_dv = DataValidation(type='list', formula1='"Must,Should,Could"', allow_blank=True)
        priority_dv.add(f'K2:K{max_row}')
        ws.add_data_validation(priority_dv)

        area_dv = _dv_list(area_names)
        if area_dv:
            area_dv.add(f'C2:C{max_row}')
            ws.add_data_validation(area_dv)

        cat_dv = _dv_list(cat_names)
        if cat_dv:
            cat_dv.add(f'D2:D{max_row}')
            ws.add_data_validation(cat_dv)

        sub_dv = _dv_list(sub_names)
        if sub_dv:
            sub_dv.add(f'E2:E{max_row}')
            ws.add_data_validation(sub_dv)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='CPMS_Import_Template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/test-suite')
def export_test_suite():
    """Export the full test suite (all categories & test cases) to Excel."""
    import openpyxl
    from io import BytesIO
    from flask import send_file
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter, quote_sheetname

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_fill_nf = PatternFill(start_color='FD7E14', end_color='FD7E14', fill_type='solid')
    lookup_fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')

    # ── Build Lookups sheet ─────────────────────────────────────────
    ws_lk = wb.create_sheet(title='Lookups')
    lk_headers = ['Area', 'Category', 'Subcategory']
    ws_lk.append(lk_headers)
    for ci in range(1, 4):
        cell = ws_lk.cell(row=1, column=ci)
        cell.font = header_font
        cell.fill = lookup_fill
        cell.alignment = Alignment(horizontal='center')

    # Collect unique areas, categories, subcategories
    all_areas = Area.query.order_by(Area.suite_type, Area.sort_order).all()
    lookup_rows = []
    for area in all_areas:
        for cat in area.categories.filter_by(parent_id=None).order_by(Category.sort_order).all():
            subs = cat.children.order_by(Category.sort_order).all()
            if subs:
                for sub in subs:
                    lookup_rows.append((area.name, cat.name, sub.name))
            else:
                lookup_rows.append((area.name, cat.name, ''))
    # Also add orphan categories (no area)
    orphan_cats = Category.query.filter_by(area_id=None, parent_id=None).order_by(Category.sort_order).all()
    for cat in orphan_cats:
        subs = cat.children.order_by(Category.sort_order).all()
        if subs:
            for sub in subs:
                lookup_rows.append(('', cat.name, sub.name))
        else:
            lookup_rows.append(('', cat.name, ''))

    for row in lookup_rows:
        ws_lk.append(list(row))

    # Auto-size lookup columns
    for col in ws_lk.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws_lk.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Named range references for data validation
    area_names = sorted(set(r[0] for r in lookup_rows if r[0]))
    cat_names = sorted(set(r[1] for r in lookup_rows if r[1]))
    sub_names = sorted(set(r[2] for r in lookup_rows if r[2]))

    def _dv_list(values, max_len=255):
        """Create a DataValidation from a list of values (falls back to no validation if too long)."""
        formula = '"' + ','.join(values) + '"'
        if len(formula) <= max_len:
            return DataValidation(type='list', formula1=formula, allow_blank=True,
                                 showErrorMessage=True, errorTitle='Invalid value',
                                 error='Please select a value from the list.')
        return None

    # ── Build test suite sheets ─────────────────────────────────────
    for suite_type in ['Functional', 'Non-Functional']:
        ws = wb.create_sheet(title=f'{suite_type} Test Suite')
        headers = ['Tier', 'Area', 'Category', 'Subcategory',
                   'Capability / Feature', 'Test Scenario', 'Pass Criteria',
                   'Evidence Required', 'Test Method', 'Priority', 'Weight']
        ws.append(headers)
        fill = header_fill if suite_type == 'Functional' else header_fill_nf
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')

        tests = TestCase.query.filter_by(suite_type=suite_type).order_by(TestCase.sort_order).all()
        for tc in tests:
            cat = Category.query.get(tc.category_id)
            root = cat
            while root.parent_id:
                root = Category.query.get(root.parent_id)
            area = Area.query.get(root.area_id) if root.area_id else None
            area_name = area.name if area else ''
            ws.append([
                tc.tier, area_name, root.name,
                tc.subcategory or '', tc.capability,
                tc.test_scenario or '', tc.pass_criteria or '',
                tc.evidence_required or '', tc.test_method or '',
                tc.priority, tc.weight,
            ])

        max_row = max(len(tests) + 1, 500)  # validate up to 500 rows

        # Add data validations
        tier_dv_copy = DataValidation(type='list', formula1='"Core,Extended"', allow_blank=True)
        tier_dv_copy.add(f'A2:A{max_row}')
        ws.add_data_validation(tier_dv_copy)

        priority_dv_copy = DataValidation(type='list', formula1='"Must,Should,Could"', allow_blank=True)
        priority_dv_copy.add(f'J2:J{max_row}')
        ws.add_data_validation(priority_dv_copy)

        area_dv = _dv_list(area_names)
        if area_dv:
            area_dv.add(f'B2:B{max_row}')
            ws.add_data_validation(area_dv)

        cat_dv = _dv_list(cat_names)
        if cat_dv:
            cat_dv.add(f'C2:C{max_row}')
            ws.add_data_validation(cat_dv)

        sub_dv = _dv_list(sub_names)
        if sub_dv:
            sub_dv.add(f'D2:D{max_row}')
            ws.add_data_validation(sub_dv)

        # Auto-size columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='CPMS_Test_Suite_Export.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/projects/<int:project_id>/export')
def project_export(project_id):
    """Export full project results (all vendors) to a single Excel workbook."""
    import openpyxl
    from io import BytesIO
    from flask import send_file
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    project = Project.query.get_or_404(project_id)
    vendors = project.vendors.all()
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    thin_border = Border(
        bottom=Side(style='thin', color='DDDDDD'),
    )

    # ── Summary sheet ────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    ws_sum.append(['CPMS Evaluation Project Report'])
    ws_sum['A1'].font = Font(bold=True, size=14)
    ws_sum.append([])
    ws_sum.append(['Project', project.name])
    ws_sum.append(['Description', project.description or ''])
    ws_sum.append(['Status', project.status])
    ws_sum.append(['Vendors Evaluated', len(vendors)])
    ws_sum.append(['Total Test Cases', TestCase.query.count()])
    ws_sum.append([])

    # Vendor summary table
    sum_headers = ['Vendor', 'Method', 'Tests', 'Scored', 'Progress %', 'Avg Score %']
    ws_sum.append(sum_headers)
    for col_idx in range(1, len(sum_headers) + 1):
        cell = ws_sum.cell(row=ws_sum.max_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    score_map = _scoring_map()
    max_score_val = max((v for v in score_map.values() if v is not None), default=5)

    for v in vendors:
        total = v.results.count()
        scored = v.results.filter(TestResult.score.isnot(None)).count() if total else 0
        total_ws = db.session.query(db.func.sum(TestResult.weighted_score)).filter(
            TestResult.vendor_id == v.id, TestResult.score.isnot(None)
        ).scalar() or 0
        total_weight = db.session.query(db.func.sum(TestCase.weight)).join(TestResult).filter(
            TestResult.vendor_id == v.id, TestResult.score.isnot(None)
        ).scalar() or 0
        avg_pct = round((total_ws / (total_weight * max_score_val) * 100) if total_weight else 0, 1)
        progress = round((scored / total * 100) if total else 0, 1)
        ws_sum.append([v.name, v.eval_method, total, scored, progress, avg_pct])

    # ── Per-vendor result sheets ─────────────────────────────────────
    for v in vendors:
        # Truncate vendor name for sheet title (max 31 chars)
        sheet_name = v.name[:28] + '...' if len(v.name) > 31 else v.name
        ws = wb.create_sheet(title=sheet_name)
        headers = ['Test ID', 'Tier', 'Area', 'Category', 'Subcategory', 'Capability',
                   'Priority', 'Weight', 'Mandatory', 'Support Level', 'Score',
                   'Weighted Score', 'Status', 'Pass/Fail', 'Notes', 'Evidence Count']
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        results = db.session.query(TestResult, TestCase).join(TestCase).filter(
            TestResult.vendor_id == v.id,
        ).order_by(TestCase.suite_type, TestCase.sort_order).all()

        for tr, tc in results:
            cat = Category.query.get(tc.category_id)
            root = cat
            while root.parent_id:
                root = Category.query.get(root.parent_id)
            area = Area.query.get(root.area_id) if root.area_id else None
            area_name = area.name if area else ''
            ev_count = Evidence.query.filter_by(test_result_id=tr.id).count()
            ws.append([
                tc.test_id_code, tc.tier, area_name, root.name, tc.subcategory or '',
                tc.capability, tc.priority, tc.weight,
                'Yes' if tc.is_mandatory else '',
                tr.support_level or '', tr.score if tr.score is not None else '',
                tr.weighted_score if tr.weighted_score is not None else '',
                tr.status, tr.pass_fail or '', tr.notes or '', ev_count,
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    # ── Comparison sheet ─────────────────────────────────────────────
    if len(vendors) > 1:
        ws_cmp = wb.create_sheet(title='Vendor Comparison')
        cmp_headers = ['Category'] + [v.name for v in vendors]
        ws_cmp.append(cmp_headers)
        for col_idx in range(1, len(cmp_headers) + 1):
            cell = ws_cmp.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        areas = Area.query.order_by(Area.suite_type, Area.sort_order).all()
        for area in areas:
            cat_ids = []
            for c in area.categories.all():
                cat_ids.append(c.id)
                for sub in c.children.all():
                    cat_ids.append(sub.id)
            if not cat_ids:
                continue
            row = [f'{area.name} ({area.suite_type})']
            for v in vendors:
                res = db.session.query(TestResult, TestCase).join(TestCase).filter(
                    TestResult.vendor_id == v.id,
                    TestCase.category_id.in_(cat_ids),
                    TestResult.score.isnot(None),
                ).all()
                total_w = sum(tc.weight for _, tc in res)
                total_ws = sum((tr.weighted_score or 0) for tr, _ in res)
                pct = round((total_ws / (total_w * max_score_val) * 100) if total_w else 0, 1)
                row.append(f'{pct}%')
            ws_cmp.append(row)

        # Totals row
        row = ['OVERALL']
        for v in vendors:
            total_ws_val = db.session.query(db.func.sum(TestResult.weighted_score)).filter(
                TestResult.vendor_id == v.id, TestResult.score.isnot(None)
            ).scalar() or 0
            total_w_val = db.session.query(db.func.sum(TestCase.weight)).join(TestResult).filter(
                TestResult.vendor_id == v.id, TestResult.score.isnot(None)
            ).scalar() or 0
            pct = round((total_ws_val / (total_w_val * max_score_val) * 100) if total_w_val else 0, 1)
            row.append(f'{pct}%')
        ws_cmp.append(row)
        ws_cmp.cell(row=ws_cmp.max_row, column=1).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'{project.name}_Full_Report.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/vendors/<int:vendor_id>/export')
def vendor_export(vendor_id):
    """Export vendor results to Excel."""
    import openpyxl
    from io import BytesIO
    from flask import send_file

    vendor = Vendor.query.get_or_404(vendor_id)
    wb = openpyxl.Workbook()

    for suite_type in ['Functional', 'Non-Functional']:
        ws = wb.create_sheet(title=f'{suite_type} Results')
        ws.append(['Test ID', 'Tier', 'Area', 'Category', 'Subcategory', 'Capability',
                    'Priority', 'Weight', 'Support Level', 'Score', 'Weighted Score',
                    'Status', 'Pass/Fail', 'Notes'])

        results = db.session.query(TestResult, TestCase).join(TestCase).filter(
            TestResult.vendor_id == vendor_id,
            TestCase.suite_type == suite_type,
        ).order_by(TestCase.sort_order).all()

        for tr, tc in results:
            cat = Category.query.get(tc.category_id)
            root = cat
            while root.parent_id:
                root = Category.query.get(root.parent_id)
            area = Area.query.get(root.area_id) if root.area_id else None
            area_name = area.name if area else ''
            ws.append([
                tc.test_id_code, tc.tier, area_name, root.name, tc.subcategory or '',
                tc.capability, tc.priority, tc.weight,
                tr.support_level or '', tr.score if tr.score is not None else '',
                tr.weighted_score if tr.weighted_score is not None else '',
                tr.status, tr.pass_fail or '', tr.notes or '',
            ])

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'{vendor.name}_evaluation_results.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════════════════
#  RUN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print("\n  CPMS Evaluation Platform")
    print("  ========================")
    print("  Open http://127.0.0.1:5000 in your browser\n")
    app.run(debug=True, host='127.0.0.1', port=5000)
