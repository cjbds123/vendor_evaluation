"""
Import the CPMS_Evaluation_Test_Suite.xlsx into the database.
Handles both Functional and Non-Functional sheets.
"""
import os
import openpyxl
from models import db, ScoringLevel, Area, Category, TestCase


# ── Default scoring scale (from Lookups sheet) ──────────────────────────────
DEFAULT_SCORING = [
    ('OOB (out-of-box)', 5, 0),
    ('Configurable', 4, 1),
    ('Custom / Professional Services', 2, 2),
    ('Roadmap', 1, 3),
    ('Not supported', 0, 4),
    ('N/A', None, 5),
]


def seed_scoring_levels():
    """Create default scoring levels if none exist."""
    if ScoringLevel.query.count() == 0:
        for label, value, order in DEFAULT_SCORING:
            db.session.add(ScoringLevel(label=label, value=value, sort_order=order))
        db.session.commit()


def _get_or_create_area(name, suite_type, sort_order=0, test_suite_id=None):
    """Find existing area or create new one."""
    q = Area.query.filter_by(name=name, suite_type=suite_type)
    if test_suite_id:
        q = q.filter_by(test_suite_id=test_suite_id)
    area = q.first()
    if not area:
        area = Area(name=name, suite_type=suite_type, sort_order=sort_order, test_suite_id=test_suite_id)
        db.session.add(area)
        db.session.flush()
    return area


def _get_or_create_category(name, suite_type, area=None, parent=None, sort_order=0):
    """Find existing category or create new one."""
    cat = Category.query.filter_by(
        name=name, suite_type=suite_type,
        area_id=area.id if area else None,
        parent_id=parent.id if parent else None,
    ).first()
    if not cat:
        cat = Category(
            name=name,
            suite_type=suite_type,
            area_id=area.id if area else None,
            parent_id=parent.id if parent else None,
            sort_order=sort_order,
        )
        db.session.add(cat)
        db.session.flush()
    return cat


def generate_test_id(suite_type):
    """Auto-generate a unique test_id_code like F-001 or NF-001."""
    prefix = 'F' if suite_type == 'Functional' else 'NF'
    last = TestCase.query.filter_by(suite_type=suite_type).order_by(TestCase.id.desc()).first()
    if last and last.test_id_code:
        try:
            num = int(last.test_id_code.split('-')[-1]) + 1
        except (ValueError, IndexError):
            num = TestCase.query.filter_by(suite_type=suite_type).count() + 1
    else:
        num = 1
    test_id_code = f'{prefix}-{num:03d}'
    while TestCase.query.filter_by(test_id_code=test_id_code).first():
        num += 1
        test_id_code = f'{prefix}-{num:03d}'
    return test_id_code


def import_excel(filepath, test_suite_id=None):
    """Import test cases from the CPMS Excel workbook.

    Returns dict with counts and any warnings.
    """
    if not os.path.exists(filepath):
        return {'error': f'File not found: {filepath}'}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    stats = {'functional': 0, 'non_functional': 0, 'areas': 0, 'categories': 0, 'warnings': [], 'test_ids': []}

    sheets = [
        ('Functional Test Suite', 'Functional'),
        ('Non-Functional Test Suite', 'Non-Functional'),
    ]

    cat_sort_counter = {}

    for sheet_name, suite_type in sheets:
        if sheet_name not in wb.sheetnames:
            stats['warnings'].append(f'Sheet "{sheet_name}" not found – skipped.')
            continue

        # Create one default Area per suite type.
        # Areas are a hierarchy level *above* Category; the Excel does not
        # contain Area information, so we create a sensible default.
        default_area = _get_or_create_area(
            f'{suite_type} Tests', suite_type, sort_order=1, test_suite_id=test_suite_id
        )
        stats['areas'] += 1

        ws = wb[sheet_name]

        # Detect whether the first column is a legacy "Test ID" column
        header_cells = [str(c.value or '').strip().lower() for c in ws[1]]
        col_offset = 1 if header_cells and header_cells[0] in ('test id', 'test_id', 'id') else 0

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

        for row_idx, row in enumerate(rows, start=2):
            if not row:  # skip empty rows
                continue

            o = col_offset  # column offset for legacy files that include Test ID
            tier            = str(row[o+0] or 'Core').strip() if len(row) > o+0 else 'Core'
            area_name       = str(row[o+1] or '').strip() if len(row) > o+1 else ''     # Area
            cat_name        = str(row[o+2] or 'Uncategorized').strip() if len(row) > o+2 else 'Uncategorized'  # Category
            subcat_name     = str(row[o+3]).strip() if len(row) > o+3 and row[o+3] else None  # Subcategory
            capability      = str(row[o+4] or '').strip() if len(row) > o+4 else ''
            test_scenario   = str(row[o+5] or '').strip() if len(row) > o+5 else ''
            pass_criteria   = str(row[o+6] or '').strip() if len(row) > o+6 else ''
            evidence_req    = str(row[o+7] or '').strip() if len(row) > o+7 else ''
            test_method     = str(row[o+8] or '').strip() if len(row) > o+8 else ''
            priority_raw    = str(row[o+9] or 'Should').strip() if len(row) > o+9 else 'Should'
            weight_raw      = row[o+10] if len(row) > o+10 else None

            # Skip rows with no meaningful content
            if not tier and not capability and not cat_name:
                continue


            # Parse weight
            try:
                weight = float(weight_raw) if weight_raw else 1.0
            except (ValueError, TypeError):
                weight = 1.0

            # Determine if mandatory (Must priority)
            is_mandatory = priority_raw.lower() == 'must'

            # Resolve area: use Area column if provided, otherwise default
            if area_name:
                area = _get_or_create_area(area_name, suite_type, sort_order=0, test_suite_id=test_suite_id)
            else:
                area = default_area

            # Get or create Category (top-level, under the resolved area)
            cat_key = (suite_type, 'top')
            if cat_key not in cat_sort_counter:
                cat_sort_counter[cat_key] = 0
            if not Category.query.filter_by(name=cat_name, suite_type=suite_type, area_id=area.id, parent_id=None).first():
                cat_sort_counter[cat_key] += 1
                stats['categories'] += 1
            category = _get_or_create_category(cat_name, suite_type, area=area, sort_order=cat_sort_counter[cat_key])

            # Get or create Subcategory (child of category) if present
            if subcat_name:
                sub_key = (suite_type, category.id)
                if sub_key not in cat_sort_counter:
                    cat_sort_counter[sub_key] = 0
                if not Category.query.filter_by(name=subcat_name, suite_type=suite_type, parent_id=category.id).first():
                    cat_sort_counter[sub_key] += 1
                subcategory = _get_or_create_category(subcat_name, suite_type, area=None, parent=category, sort_order=cat_sort_counter[sub_key])
                assign_category_id = subcategory.id
            else:
                assign_category_id = category.id
                subcat_name = None

            # Auto-generate unique test ID
            test_id_code = generate_test_id(suite_type)

            tc = TestCase(
                test_id_code=test_id_code,
                tier=tier,
                category_id=assign_category_id,
                subcategory=subcat_name,
                capability=capability,
                test_scenario=test_scenario,
                pass_criteria=pass_criteria,
                evidence_required=evidence_req,
                test_method=test_method,
                priority=priority_raw,
                weight=weight,
                is_mandatory=is_mandatory,
                suite_type=suite_type,
                sort_order=row_idx,
            )
            db.session.add(tc)
            db.session.flush()  # get tc.id
            stats['test_ids'].append(tc.id)

            if suite_type == 'Functional':
                stats['functional'] += 1
            else:
                stats['non_functional'] += 1

    db.session.commit()
    return stats
